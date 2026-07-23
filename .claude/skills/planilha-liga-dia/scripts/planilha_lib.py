"""
planilha_lib.py — motor compartilhado da skill planilha-liga-dia (Copa /
Série A / Série B), que consolida o que antes eram 3 skills quase-idênticas
(copa-planilha-dia, serie-a-planilha-dia, serie-b-planilha-dia).

Uso típico (ver SKILL.md para o passo a passo completo):

    import sys
    sys.path.insert(0, "/mnt/skills/user/planilha-liga-dia/scripts")
    from planilha_lib import *
    from ligas import estilo_db_path, output_path_for   # camada por liga

    df = load_all_matches()  # todos os CSVs de partidas em data/matches/*.csv
    hist_A = get_historico("Ceará", df)
    hist_B = get_historico("Athletic Club", df)
    attach_estilo(hist_A, estilo_db_path("serieb"))
    attach_estilo(hist_B, estilo_db_path("serieb"))

    games = [
        dict(teamA="Ceará", teamB="Athletic Club", hist_A=hist_A, hist_B=hist_B,
             jdd_A=dict(...), jdd_B=dict(...), mercados_rows=[...], fontes=dict(...))
    ]
    build_workbook(games, template_path=".../Copa_Template_Simplificado.xlsx",
                   output_path="/mnt/user-data/outputs/" + output_path_for("serieb", "13jul"),
                   data_jogo="13/07/2026")

Este módulo cuida da parte MECÂNICA (dados históricos dos CSVs, clonagem de
fórmulas, redimensionamento das 3 tabelas do Excel, ajuste de mando, fórmula
Pró/Contra robusta) — é 100% liga-agnóstico. O que MUDA por competição
(CSVs agregados de liga/jogador, banco de estilos, se é torneio neutro) vive
em ligas.py. A parte de JULGAMENTO (quais jogos, odds do dia, estilo,
probabilidades das fontes) continua sendo feita pelo Claude a cada sessão —
ver SKILL.md.
"""
import os
import re
import json
import glob
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))

# Todos os CSVs de PARTIDAS (histórico jogo-a-jogo) ficam juntos aqui,
# independente da liga — um time promovido/rebaixado carrega histórico dos
# dois lados (ver briefing_serieb.md), e times de seleção (Copa) referenciam
# vários arquivos de competição (worldcup/friendlies/qualifiers). Quem
# desambigua por liga é get_historico() (filtra por nome de time), não o
# glob de carregamento.
CSV_GLOB_DEFAULT = os.path.join(_HERE, "..", "data", "matches", "*.csv")
ESTILO_DB_PATH = os.path.join(_HERE, "..", "data", "estilos_selecoes.json")  # default: Copa

# Alias de nomes divergentes entre CSV do FootyStats e nomenclatura usada nas
# planilhas / casas de aposta. Isso aqui é só o default (Copa); cada liga tem
# o seu próprio em ligas.LIGAS[liga].csv_alias — passe alias= explicitamente
# em get_historico() quando for time de liga diferente da Copa.
CSV_ALIAS = {
    "USMNT": "USA",       # CSV usa "USA", planilha usa "USMNT"
    "Congo DR": "DR Congo",
}


# ---------------------------------------------------------------------------
# 1. DADOS HISTÓRICOS (Times sheet)
# ---------------------------------------------------------------------------

def load_all_matches(csv_glob=CSV_GLOB_DEFAULT):
    """Carrega e concatena todos os CSVs do FootyStats em /mnt/project/."""
    dfs = []
    for f in glob.glob(csv_glob):
        try:
            d = pd.read_csv(f)
            d["__src"] = f.split("/")[-1]
            dfs.append(d)
        except Exception as e:
            print("ERRO ao ler", f, e)
    df = pd.concat(dfs, ignore_index=True, sort=False)
    return df[df["status"] == "complete"]


# ---------------------------------------------------------------------------
# 1b. MINERAÇÃO DE PADRÃO ESTRUTURAL (independente de odd) — 16/07
# ---------------------------------------------------------------------------
# Acha limiares que batem 100% (ou 0%) do tempo por (time, mando), na
# temporada COMPLETA de uma liga (não só os N jogos recentes de um time do
# dia). Serve pra levantar candidatos a investigar — NUNCA pra apostar
# direto: um padrão 100% em 7-10 jogos é indício, não prova, e pode quebrar
# com o próximo jogo (testado na prática: Botafogo Casa Cartões Total "sempre
# >3.5" caiu de 100% pra ~86% ao trocar de 6 pra 7 jogos na amostra). Sem
# condicionar por aderência de estilo/favoritismo, qualquer achado aqui
# provavelmente já está precificado pelo mercado -- usar como ponto de
# partida pra investigação, não como sinal pronto.

PADROES_THRESHOLDS_DEFAULT = {
    "gols_total": 2.5, "gols_pro": 1.5,
    "cart_total": 3.5, "cart_pro": 2.5,
    "esc_total": 8.5, "esc_pro": 3.5,
    "chutes_total": 22.5, "chutesgol_pro": 3.5,
}


def _achatar_casa_fora(df):
    """Transforma o df 'largo' do FootyStats (1 linha = 1 jogo, colunas
    home_*/away_*) em formato 'longo' (1 linha = 1 time em 1 jogo), igual à
    aba Times. Usado só para mineração de padrão -- não mexe no fluxo normal
    de get_historico/attach_estilo."""
    req = ["home_team_name", "away_team_name", "date_GMT",
           "home_team_goal_count", "away_team_goal_count",
           "home_team_corner_count", "away_team_corner_count",
           "home_team_yellow_cards", "home_team_red_cards",
           "away_team_yellow_cards", "away_team_red_cards",
           "home_team_shots", "away_team_shots",
           "home_team_shots_on_target", "away_team_shots_on_target"]
    faltando = [c for c in req if c not in df.columns]
    if faltando:
        raise ValueError(f"CSV sem colunas esperadas: {faltando}")

    casa = pd.DataFrame(dict(
        time=df.home_team_name, adv=df.away_team_name, mando="Casa", data=df.date_GMT,
        gols_pro=df.home_team_goal_count, gols_con=df.away_team_goal_count,
        esc_pro=df.home_team_corner_count, esc_con=df.away_team_corner_count,
        cart_pro=df.home_team_yellow_cards + df.home_team_red_cards,
        cart_con=df.away_team_yellow_cards + df.away_team_red_cards,
        chutes_pro=df.home_team_shots, chutes_con=df.away_team_shots,
        chutesgol_pro=df.home_team_shots_on_target, chutesgol_con=df.away_team_shots_on_target,
    ))
    fora = pd.DataFrame(dict(
        time=df.away_team_name, adv=df.home_team_name, mando="Fora", data=df.date_GMT,
        gols_pro=df.away_team_goal_count, gols_con=df.home_team_goal_count,
        esc_pro=df.away_team_corner_count, esc_con=df.home_team_corner_count,
        cart_pro=df.away_team_yellow_cards + df.away_team_red_cards,
        cart_con=df.home_team_yellow_cards + df.home_team_red_cards,
        chutes_pro=df.away_team_shots, chutes_con=df.home_team_shots,
        chutesgol_pro=df.away_team_shots_on_target, chutesgol_con=df.home_team_shots_on_target,
    ))
    full = pd.concat([casa, fora], ignore_index=True)
    full["esc_total"] = full.esc_pro + full.esc_con
    full["cart_total"] = full.cart_pro + full.cart_con
    full["gols_total"] = full.gols_pro + full.gols_con
    full["chutes_total"] = full.chutes_pro + full.chutes_con
    return full


def buscar_padroes_liga(df=None, csv_glob=None, min_jogos=7, thresholds=None, teams=None, min_pct=1.0):
    """Varre TODOS os jogos completos de uma liga (não só os N recentes de um
    time específico) procurando limiares que batem >=min_pct (ou <=1-min_pct)
    do tempo por (time, mando).

    df        : DataFrame já carregado via load_all_matches (para restringir
                a um CSV específico, ex. só brazilseriea*.csv). Se None, usa
                csv_glob.
    csv_glob  : glob pattern alternativo (ex. "/mnt/project/brazilseriea*.csv"),
                só usado se df não for passado.
    min_jogos : amostra mínima por (time, mando) para o achado contar.
    thresholds: dict {metrica: limiar}. Default cobre gols/cartões/escanteios/
                chutes, pró e total. Métricas disponíveis: gols_pro,
                gols_total, cart_pro, cart_total, esc_pro, esc_total,
                chutes_total, chutesgol_pro.
    teams     : lista opcional para restringir a alguns times.
    min_pct   : confiança mínima do padrão (padrão 1.0 = só 100%/0% exato,
                igual ao comportamento original). Passar min_pct=0.8, por
                exemplo, pra flexibilizar e pegar padrões que batem >=80% do
                tempo (ou <=20%, na direção "abaixo") -- exigir 100% numa
                amostra de 7-18 jogos é frágil (um jogo a mais já derrubou o
                padrão do Botafogo em cartões, visto na prática em 16/07).

    Retorna DataFrame com colunas: time, mando, metrica, limiar, direcao
    ('acima'/'abaixo'), pct (0-100), n, min, max, media -- ordenado por n
    desc. `media` ajuda a distinguir um 88% "confortável" (média bem acima
    do limiar) de um 88% "no fio da navalha" (média mal passa do limiar).

    🚨 Achado aqui é candidato a investigar, não sinal pronto -- sem
    condicionar por aderência/favoritismo, o mercado provavelmente já
    precifica. E com n=7-10 (uma temporada só), um jogo a mais pode quebrar
    o padrão -- ver nota acima."""
    if df is None:
        if csv_glob is None:
            raise ValueError("Passe df ou csv_glob.")
        df = load_all_matches(csv_glob=csv_glob)

    thresholds = thresholds or PADROES_THRESHOLDS_DEFAULT
    long = _achatar_casa_fora(df)
    if teams:
        long = long[long.time.isin(teams)]

    achados = []
    for (time, mando), g in long.groupby(["time", "mando"]):
        if len(g) < min_jogos:
            continue
        for metrica, limiar in thresholds.items():
            if metrica not in g:
                continue
            pct = (g[metrica] > limiar).mean()
            if pct >= min_pct or pct <= (1 - min_pct):
                achados.append(dict(
                    time=time, mando=mando, metrica=metrica, limiar=limiar,
                    direcao="acima" if pct >= min_pct else "abaixo",
                    pct=round(pct * 100, 1),
                    n=len(g), min=g[metrica].min(), max=g[metrica].max(),
                    media=round(g[metrica].mean(), 2),
                ))
    res = pd.DataFrame(achados)
    if len(res):
        res = res.sort_values(["n", "pct"], ascending=[False, False]).reset_index(drop=True)
    return res


def montar_aba_padroes(xlsx_path, games, csv_glob, min_pct=0.8, min_jogos=7, thresholds=None):
    """Cria (ou substitui) a aba 'Padrões' na planilha do dia, com os achados
    de buscar_padroes_liga() já filtrados para o lado (mando) que cada time
    efetivamente joga hoje -- o time A joga em casa, então só entra o padrão
    'Casa' dele; o time B joga fora, só entra o padrão 'Fora' dele.

    xlsx_path : planilha do dia já montada (roda depois de build_workbook).
    games     : mesma lista de dicts passada pro build_workbook (usa só
                teamA/teamB daqui).
    csv_glob  : CSV(s) da liga pra minerar o padrão (ex.
                "/mnt/project/brazilseriea*.csv").
    min_pct/min_jogos/thresholds: repassados pra buscar_padroes_liga().

    Se não houver NENHUM achado para os times do dia, cria a aba mesmo assim
    com uma linha avisando isso (nunca deixa a aba ausente sem explicação)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pares = []  # (time, mando_que_joga_hoje, jogo_label)
    for g in games:
        jogo = f"{g['teamA']} x {g['teamB']}"
        pares.append((g["teamA"], "Casa", jogo))
        pares.append((g["teamB"], "Fora", jogo))

    times_hoje = list({p[0] for p in pares})
    res = buscar_padroes_liga(csv_glob=csv_glob, min_jogos=min_jogos,
                               thresholds=thresholds, teams=times_hoje, min_pct=min_pct)

    if len(res):
        mando_por_time = {t: m for t, m, _j in pares}
        jogo_por_time = {t: j for t, _m, j in pares}
        res = res[res.apply(lambda r: mando_por_time.get(r["time"]) == r["mando"], axis=1)].copy()
        res["jogo"] = res["time"].map(jogo_por_time)
        res = res.sort_values(["jogo", "pct"], ascending=[True, False]).reset_index(drop=True)

    wb = load_workbook(xlsx_path)
    if "Padrões" in wb.sheetnames:
        del wb["Padrões"]
    ws = wb.create_sheet("Padrões")

    headers = ["Jogo", "Time", "Mando", "Métrica", "Limiar", "Média", "Margem", "%", "Direção", "N jogos", "Mín", "Máx"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    if len(res) == 0:
        ws.cell(2, 1, f"Nenhum padrão >= {min_pct*100:.0f}% encontrado para os times de hoje "
                       f"(min. {min_jogos} jogos por time/mando).")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    else:
        for i, row in res.iterrows():
            r = i + 2
            margem = round(row["media"] - row["limiar"], 2)
            if row["direcao"] == "abaixo":
                margem = round(row["limiar"] - row["media"], 2)  # positivo = confortável nos dois sentidos
            ws.cell(r, 1, row["jogo"])
            ws.cell(r, 2, row["time"])
            ws.cell(r, 3, row["mando"])
            ws.cell(r, 4, row["metrica"])
            ws.cell(r, 5, row["limiar"])
            ws.cell(r, 6, row["media"])
            ws.cell(r, 7, margem)
            ws.cell(r, 8, f'{row["pct"]:.0f}%')
            ws.cell(r, 9, row["direcao"])
            ws.cell(r, 10, row["n"])
            ws.cell(r, 11, row["min"])
            ws.cell(r, 12, row["max"])
            # destaque: precisa de % alto E margem confortável (média não "no fio")
            if row["pct"] >= 95 and abs(margem) >= 1:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                for c in range(1, len(headers) + 1):
                    ws.cell(r, c).fill = fill
            elif abs(margem) < 0.5:
                # % pode estar alto, mas a média está "no fio da navalha" -- sinalizar em amarelo
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for c in range(1, len(headers) + 1):
                    ws.cell(r, c).fill = fill

        aviso_row = len(res) + 3
        ws.cell(aviso_row, 1,
                "⚠️ Sem condicionar por aderência de estilo/favoritismo — o mercado provavelmente já "
                "precifica isso. Amostra de 1 temporada (7-18 jogos): tratar como candidato a "
                "investigar, não sinal pronto para apostar.")
        ws.merge_cells(start_row=aviso_row, start_column=1, end_row=aviso_row, end_column=len(headers))
        ws.cell(aviso_row, 1).font = Font(italic=True, size=9)

    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 4)
    ws.column_dimensions["A"].width = 26

    wb.save(xlsx_path)
    return xlsx_path


def _comp_name(src):
    """Mapeia nome de arquivo CSV -> (nome da competição, mando forçado ou None).
    Mando forçado 'Neutro' para Copa do Mundo e Eurocopa (jogos em local
    neutro na convenção usada nesta planilha)."""
    s = src.lower()
    if "worldcup" in s:
        return ("Copa do Mundo 2026", "Neutro")
    if "eurochampionship" in s:
        return ("Eurocopa 2024", "Neutro")
    if "nationsleague" in s:
        return ("Nations League", None)
    if "friendliesmatches2026" in s:
        return ("Amistoso/Preparação 2026", None)
    if "friendliesmatches2025" in s:
        return ("Amistoso Internacional", None)
    if "qualificationeurope" in s:
        return ("Elim. Copa (Europa)", None)
    if "qualificationafrica" in s:
        return ("Elim. Copa (África)", None)
    if "qualificationconcacaf" in s:
        return ("Elim. Copa (CONCACAF)", None)
    if "qualificationasia" in s:
        return ("Elim. Copa (Ásia)", None)
    if "qualificationsouthamerica" in s:
        return ("Elim. Copa (Am. do Sul)", None)
    if "qualificationintercontinental" in s:
        return ("Repescagem Intercontinental", "Neutro")
    return (src, None)


def get_historico(team, df, n=None, alias=None):
    """Retorna os jogos completos de `team`, mais recentes primeiro, já na
    perspectiva do time (Pró/Contra, mando, favoritismo normalizado).

    n: quantidade de jogos a retornar. PADRÃO (None) = TODOS os jogos
       completos encontrados no CSV pra esse time (mudou de 15 fixo pra
       'todos' em 16/07 -- ver PROTOCOLO/histórico de decisão: com só 15
       jogos, um padrão que parecia 100% virava ~86% ao entrar mais um jogo
       na amostra; usar a temporada inteira reduz esse risco de amostra
       pequena). Passar um número (ex. n=15) se quiser limitar de propósito
       (raro -- só faz sentido se o time trocou de técnico/sistema recente e
       você quer isolar só a fase nova, aí normalmente é melhor filtrar por
       data em vez de por contagem).

    `alias`: nome alternativo a procurar no CSV se `team` não bater
    (ex.: get_historico("USMNT", df, alias="USA")).
    Se `n` for passado explicitamente e vier menos jogos que isso, avisa
    (Regra 2 do protocolo -- reduzir stake/qualificar a análise).
    """
    lookup = alias or team
    sub = df[(df["home_team_name"] == lookup) | (df["away_team_name"] == lookup)].copy()
    sub = sub.sort_values("timestamp", ascending=False)
    if n is not None:
        sub = sub.head(n)
        if len(sub) < n:
            print(f"⚠️  {team}: apenas {len(sub)}/{n} jogos encontrados no CSV — "
                  f"considerar amostra parcial (Regra 2 do protocolo).")
    else:
        print(f"ℹ️  {team}: {len(sub)} jogos completos encontrados (histórico completo, sem limite).")
    out = []
    for _, r in sub.iterrows():
        is_home = r["home_team_name"] == lookup
        opp = r["away_team_name"] if is_home else r["home_team_name"]
        d = datetime.datetime.strptime(r["date_GMT"].split(" - ")[0], "%b %d %Y")
        comp, forced_mando = _comp_name(r["__src"])
        mando = forced_mando if forced_mando else ("Casa" if is_home else "Fora")

        def pick(a, b):
            return a if is_home else b

        gf, ga = pick(r["home_team_goal_count"], r["away_team_goal_count"]), pick(r["away_team_goal_count"], r["home_team_goal_count"])
        xgf, xga = pick(r["team_a_xg"], r["team_b_xg"]), pick(r["team_b_xg"], r["team_a_xg"])
        cf, ca = pick(r["home_team_corner_count"], r["away_team_corner_count"]), pick(r["away_team_corner_count"], r["home_team_corner_count"])
        yf = pick(r["home_team_yellow_cards"] + r["home_team_red_cards"], r["away_team_yellow_cards"] + r["away_team_red_cards"])
        ya = pick(r["away_team_yellow_cards"] + r["away_team_red_cards"], r["home_team_yellow_cards"] + r["home_team_red_cards"])
        sf, sa = pick(r["home_team_shots"], r["away_team_shots"]), pick(r["away_team_shots"], r["home_team_shots"])
        sotf, sota = pick(r["home_team_shots_on_target"], r["away_team_shots_on_target"]), pick(r["away_team_shots_on_target"], r["home_team_shots_on_target"])
        posf = pick(r["home_team_possession"], r["away_team_possession"])
        ff, fa = pick(r["home_team_fouls"], r["away_team_fouls"]), pick(r["away_team_fouls"], r["home_team_fouls"])
        htf, hta = pick(r["home_team_goal_count_half_time"], r["away_team_goal_count_half_time"]), pick(r["away_team_goal_count_half_time"], r["home_team_goal_count_half_time"])
        oh, od_, oa = r["odds_ft_home_team_win"], r["odds_ft_draw"], r["odds_ft_away_team_win"]
        try:
            ph, pd_, pa = 1 / oh, 1 / od_, 1 / oa
            s = ph + pd_ + pa
            fav = (ph / s) if is_home else (pa / s)
        except Exception:
            fav = None

        # sentinela -1 (dado ausente no CSV, comum em goleadas contra seleções
        # fracas que o FootyStats não cobre em detalhe). Não deixamos -1 cru
        # na planilha (pareceria dado real negativo); usamos um placeholder
        # conservador coerente com o placar e sinalizamos para o Claude
        # revisar/ajustar antes de fechar a análise (Regra 2 do protocolo).
        missing = cf == -1
        if missing:
            print(f"⚠️  {team} x {opp} ({d.strftime('%d/%m/%Y')}): estatísticas "
                  f"ausentes no CSV (-1) além do placar. Usando placeholder "
                  f"conservador — revisar/ajustar manualmente se o jogo for "
                  f"relevante para a análise.")
            margem = gf - ga
            if margem >= 3:  # goleada
                cf, ca, yf, ya, sf, sa, sotf, sota, posf, ff, fa = 9, 1, 1, 2, 22, 3, 9, 1, 68, 9, 12
                if xgf == 0 and xga == 0:
                    xgf, xga = 3.1, 0.2
            else:
                cf, ca, yf, ya, sf, sa, sotf, sota, posf, ff, fa = 5, 4, 1, 1, 12, 10, 5, 4, 50, 10, 10
                if xgf == 0 and xga == 0:
                    xgf, xga = 1.3, 1.1

        rec = dict(
            data=d.strftime("%d/%m/%Y"), comp=comp, adv=opp, mando=mando,
            gf=int(gf), ga=int(ga), xgf=round(float(xgf), 2), xga=round(float(xga), 2),
            cf=int(cf), ca=int(ca), yf=int(yf), ya=int(ya),
            sf=int(sf), sa=int(sa), sotf=int(sotf), sota=int(sota),
            posf=int(posf), ff=int(ff), fa=int(fa), htf=int(htf), hta=int(hta),
            fav=round(fav, 3) if fav else None, missing_stats=missing,
        )
        out.append(rec)
    return out


def attach_estilo(historico, estilo_db_path=ESTILO_DB_PATH, overrides=None):
    """Preenche estilo_text/bb/pa/tr/pos/bp em cada linha do histórico usando
    o banco persistente de estilos (data/estilos_selecoes.json), com
    `overrides` (dict adversário -> tupla) tendo prioridade. Levanta erro
    listando adversários faltantes para o Claude preencher manualmente."""
    db = json.load(open(estilo_db_path, encoding="utf-8"))
    overrides = overrides or {}
    faltando = []
    for r in historico:
        if r["adv"] in overrides:
            est = overrides[r["adv"]]
            r["estilo_text"], r["bb"], r["pa"], r["tr"], r["pos"], r["bp"] = est
        elif r["adv"] in db:
            e = db[r["adv"]]
            r["estilo_text"], r["bb"], r["pa"], r["tr"], r["pos"], r["bp"] = (
                e["estilo_text"], e["bb"], e["pa"], e["tr"], e["pos"], e["bp"])
        else:
            faltando.append(r["adv"])
    if faltando:
        raise ValueError(
            "Faltam notas de estilo para: " + ", ".join(sorted(set(faltando))) +
            ". Preencha em ESTILO overrides (dict adversario -> "
            "(texto, bloco_baixo, pressao_alta, transicao, posse, bola_parada), "
            "1-5 cada) e rode attach_estilo de novo. Depois de decidir, "
            "salve no banco persistente com save_estilo_db() para reaproveitar."
        )
    return historico


def save_estilo_db(new_entries, estilo_db_path=ESTILO_DB_PATH):
    """new_entries: dict {nome_time: (texto, bb, pa, tr, pos, bp)}.
    Mescla no banco persistente para as próximas planilhas."""
    db = json.load(open(estilo_db_path, encoding="utf-8"))
    for name, (texto, bb, pa, tr, pos, bp) in new_entries.items():
        db[name] = dict(estilo_text=texto, bb=bb, pa=pa, tr=tr, pos=pos, bp=bp)
    json.dump(db, open(estilo_db_path, "w", encoding="utf-8"), indent=1, ensure_ascii=False)


# ---------------------------------------------------------------------------
# 2. CLONAGEM DE FÓRMULAS (a parte que permite redimensionar as tabelas)
# ---------------------------------------------------------------------------

def clone_formula(template_text, template_row, new_row):
    """Substitui referências de linha RELATIVAS (sem $ antes do número) que
    batem com template_row por new_row. Referências ABSOLUTAS ($linha) —
    usadas nos limites de range fixos — não são tocadas aqui (ver
    fix_range_end)."""
    pattern = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")

    def repl(m):
        col, dollar_row, row = m.groups()
        if dollar_row == "" and row == str(template_row):
            return f"{col}{new_row}"
        return m.group(0)

    return pattern.sub(repl, template_text)


def fix_range_end(text, old_end, new_end):
    """Ajusta o limite final de ranges absolutos do tipo $COL$2:$COL$old_end
    para $COL$2:$COL$new_end (usado ao redimensionar Times ou Jogos do Dia)."""
    pattern = re.compile(r"(\$[A-Z]{1,3}\$2:\$[A-Z]{1,3}\$)" + str(old_end) + r"\b")
    return pattern.sub(lambda m: m.group(1) + str(new_end), text)


def _set_formula(ws, row, col, text, is_array):
    if is_array:
        ref = f"{get_column_letter(col)}{row}"
        ws.cell(row, col, ArrayFormula(ref=ref, text=text))
    else:
        ws.cell(row, col, text)


def _capture_template(ws, row, cols):
    """Lê as fórmulas da linha-modelo `row` nas colunas `cols` (lista de
    índices). Retorna dict col -> (texto, is_array)."""
    out = {}
    for col in cols:
        v = ws.cell(row, col).value
        if v is None:
            continue
        if isinstance(v, ArrayFormula):
            out[col] = (v.text, True)
        else:
            out[col] = (v, False)
    return out


# ---------------------------------------------------------------------------
# 3. CONSTRUÇÃO DA PLANILHA
# ---------------------------------------------------------------------------

TIMES_INPUT_COLS = {
    "Time": 1, "Data": 2, "Competição": 3, "Adversário": 4, "Mando": 5,
    "Gols Pró": 6, "Gols Contra": 7, "xG Pró": 8, "xG Contra": 9,
    "Escanteios Pró": 10, "Escanteios Contra": 11, "Cartões Pró": 12, "Cartões Contra": 13,
    "Chutes Pró": 14, "Chutes Contra": 15, "Chutes Gol Pró": 16, "Chutes Gol Contra": 17,
    "Posse Pró": 18, "Faltas Time": 19, "Faltas Adv": 20, "Gols 1T Time": 21, "Gols 1T Adv": 22,
    "Favoritismo": 26, "Estilo Adv": 27, "Bloco Baixo": 28, "Pressão Alta": 29,
    "Transição": 30, "Posse/Dom": 31, "Bola Parada": 32,
}
TIMES_FORMULA_COLS = {"W": 23, "X": 24, "Y": 25, "AG": 33, "AH": 34, "AI": 35, "AK": 37}
JDD_FORMULA_COLS = list(range(15, 37))  # O..AJ
MER_FORMULA_COLS = [6, 8, 9, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21]  # F,H,I,J,L,M,O..U

MERCADOS_TEMPLATE_20 = [
    # (mercado, tipo, crit1_template, crit2_template)  {A}=TimeA {B}=TimeB
    ("BTTS Sim", "BTTS", "BTTS Sim", None),
    ("BTTS Não", "BTTS", "BTTS Não", None),
    ("CA: {FAV} DC + Under 3.5", "Combinação", "{FAVKEY} DC", "Under 3.5 | 0.84"),
    ("CA: {FAV} vence + Over 1.5", "Combinação", "{FAVKEY} vence", "Over 1.5"),
    ("CA: {FAV} vence + BTTS Não", "Combinação", "{FAVKEY} vence", "BTTS Não | 0.70"),
    ("{A} DC", "Dupla Chance", "TimeA DC", None),
    ("{B} DC", "Dupla Chance", "TimeB DC", None),
    ("Over 0.5 gols 1ºT", "Gols 1º Tempo", "Over 0.5 1T", None),
    ("Under 0.5 gols 1ºT", "Gols 1º Tempo", "Under 0.5 1T", None),
    ("{FAV} -0.5 AH (vence)", "Handicap Asiático", "{FAVKEY} AH -0.5", None),
    ("{DOG} +0.5 AH", "Handicap Asiático", "{DOGKEY} AH +0.5", None),
    ("{FAV} -1.5 AH", "Handicap Asiático", "{FAVKEY} AH -1.5", None),
    ("{A} vence", "Resultado", "TimeA vence", None),
    ("Empate", "Resultado", "Empate", None),
    ("{B} vence", "Resultado", "TimeB vence", None),
    ("Over 2.5 gols", "Total gols", "Over 2.5", None),
    ("Under 2.5 gols", "Total gols", "Under 2.5", None),
    ("Over 1.5 gols", "Total gols", "Over 1.5", None),
    ("Under 1.5 gols", "Total gols", "Under 1.5", None),
    ("Over 3.5 gols", "Total gols", "Over 3.5", None),
]


def mercados_rows_for_game(teamA, teamB, favorito, odds_and_pfont, displayA=None, displayB=None):
    """Gera as 20 linhas de Mercados para um jogo a partir do template acima.
    favorito: "A" ou "B" (qual dos dois é o favorito, define os textos das
    linhas de DC/AH combinadas).
    odds_and_pfont: lista de 20 tuplas (p_font_str, odd_mercado) NA MESMA
    ORDEM de MERCADOS_TEMPLATE_20 (pesquisar odds reais + estimar via Poisson
    onde não houver cotação direta — ver SKILL.md passo 4).
    displayA/displayB: nome em português para os rótulos de mercado (ex.:
    "França", "Marrocos"). Se omitido, usa teamA/teamB (nome do CSV) direto.
    IMPORTANTE: teamA/teamB devem ser exatamente os nomes usados em Times/
    Jogos do Dia (chave de lookup); displayA/displayB é só cosmético."""
    dA, dB = displayA or teamA, displayB or teamB
    fav_name, dog_name = (dA, dB) if favorito == "A" else (dB, dA)
    fav_key, dog_key = ("TimeA", "TimeB") if favorito == "A" else ("TimeB", "TimeA")
    jogo = f"{teamA} x {teamB}"
    rows = []
    for (mercado_t, tipo, c1_t, c2_t), (pfont, odd) in zip(MERCADOS_TEMPLATE_20, odds_and_pfont):
        mercado = mercado_t.format(A=dA, B=dB, FAV=fav_name, DOG=dog_name)
        c1 = c1_t.format(FAVKEY=fav_key, DOGKEY=dog_key) if c1_t else c1_t
        c2 = c2_t
        rows.append((jogo, mercado, tipo, c1, c2, pfont, odd))
    return rows


def build_workbook(games, template_path, output_path, data_jogo):
    """games: lista de dicts, um por partida do dia, cada um com:
        teamA, teamB           -- nomes exatamente como usados no CSV/planilha
        hist_A, hist_B         -- listas retornadas por get_historico()+attach_estilo()
        jdd_A, jdd_B           -- dict com campos de Jogos do Dia (ver SKILL.md):
            estilo_time, estilo_adv, fav, bb, pa, tr, pos, bp, ataca_fundo,
            contexto, obs
        mercados_rows          -- lista de 20 tuplas, ver mercados_rows_for_game()
        fontes                 -- dict: placar_modal, p_a, p_empate, p_b,
                                   consenso, melhor_aposta, atencao
    data_jogo: string "DD/MM/AAAA" usada na coluna Data de Jogos do Dia.

    Constrói a planilha completa (Times + Jogos do Dia + Mercados + Fontes)
    a partir do template, redimensionando as 3 tabelas Excel para o número
    exato de jogos do dia, e salva em output_path.
    """
    n_games = len(games)
    # Antes: n_games*30 assumia 15 jogos fixos por time (2 times x 15). Desde
    # que get_historico() passou a trazer TODOS os jogos por padrão (16/07),
    # isso tem que ser dinâmico -- soma o tamanho real de cada histórico.
    n_times_rows = sum(len(g["hist_A"]) + len(g["hist_B"]) for g in games)
    n_jdd_rows = n_games * 2          # 2 linhas por jogo (perspectiva A e B)
    n_mer_rows = n_games * 20         # 20 mercados por jogo

    wb = load_workbook(template_path)

    # ---------------- TIMES ----------------
    ws = wb["Times"]
    times_tpl = {name: ws.cell(2, col).value for name, col in TIMES_FORMULA_COLS.items()}
    old_last_row = ws.tables["Tabela15"].ref.split(":")[1]
    old_last_row_n = int(re.search(r"\d+", old_last_row).group())
    if old_last_row_n > n_times_rows + 1:
        ws.delete_rows(n_times_rows + 2, old_last_row_n - (n_times_rows + 1))
    elif old_last_row_n < n_times_rows + 1:
        ws.insert_rows(old_last_row_n + 1, (n_times_rows + 1) - old_last_row_n)

    def fix5to(text, new_n):
        return fix_range_end(text, 5, new_n)

    r = 2
    for g in games:
        for team_name, hist in ((g["teamA"], g["hist_A"]), (g["teamB"], g["hist_B"])):
            for rec in hist:
                for label, col in TIMES_INPUT_COLS.items():
                    key = {
                        "Time": None, "Data": "data", "Competição": "comp", "Adversário": "adv",
                        "Mando": "mando", "Gols Pró": "gf", "Gols Contra": "ga",
                        "xG Pró": "xgf", "xG Contra": "xga", "Escanteios Pró": "cf",
                        "Escanteios Contra": "ca", "Cartões Pró": "yf", "Cartões Contra": "ya",
                        "Chutes Pró": "sf", "Chutes Contra": "sa", "Chutes Gol Pró": "sotf",
                        "Chutes Gol Contra": "sota", "Posse Pró": "posf", "Faltas Time": "ff",
                        "Faltas Adv": "fa", "Gols 1T Time": "htf", "Gols 1T Adv": "hta",
                        "Favoritismo": "fav", "Estilo Adv": "estilo_text", "Bloco Baixo": "bb",
                        "Pressão Alta": "pa", "Transição": "tr", "Posse/Dom": "pos",
                        "Bola Parada": "bp",
                    }[label]
                    ws.cell(r, col, team_name if key is None else rec[key])
                for name, col in TIMES_FORMULA_COLS.items():
                    txt = clone_formula(times_tpl[name], 2, r)
                    if name in ("AG", "AH", "AI"):
                        txt = fix5to(txt, n_jdd_rows + 1)
                    ws.cell(r, col, txt)
                r += 1

    tbl = ws.tables["Tabela15"]
    tbl.ref = f"A1:AI{n_times_rows + 1}"
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref

    # ---------------- JOGOS DO DIA ----------------
    wsj = wb["Jogos do Dia"]
    jdd_tpl = _capture_template(wsj, 2, JDD_FORMULA_COLS)
    old_last = int(re.search(r"\d+", wsj.tables["Tabela26"].ref.split(":")[1]).group())
    if old_last > n_jdd_rows + 1:
        wsj.delete_rows(n_jdd_rows + 2, old_last - (n_jdd_rows + 1))
    elif old_last < n_jdd_rows + 1:
        wsj.insert_rows(old_last + 1, (n_jdd_rows + 1) - old_last)

    def fix61to(text, new_n):
        return fix_range_end(text, 61, new_n)

    r = 2
    for g in games:
        for team_name, adv_name, jdd in (
            (g["teamA"], g["teamB"], g["jdd_A"]), (g["teamB"], g["teamA"], g["jdd_B"])
        ):
            wsj.cell(r, 1, data_jogo)
            wsj.cell(r, 2, team_name)
            wsj.cell(r, 3, adv_name)
            wsj.cell(r, 4, jdd["estilo_time"])
            wsj.cell(r, 5, jdd["estilo_adv"])
            wsj.cell(r, 6, jdd["fav"])
            wsj.cell(r, 7, jdd["bb"])
            wsj.cell(r, 8, jdd["pa"])
            wsj.cell(r, 9, jdd["tr"])
            wsj.cell(r, 10, jdd["pos"])
            wsj.cell(r, 11, jdd["bp"])
            wsj.cell(r, 12, jdd["ataca_fundo"])
            wsj.cell(r, 13, jdd["contexto"])
            wsj.cell(r, 14, jdd["obs"])
            for col, (txt, is_arr) in jdd_tpl.items():
                new_txt = clone_formula(txt, 2, r)
                new_txt = fix61to(new_txt, n_times_rows + 1)
                _set_formula(wsj, r, col, new_txt, is_arr)
            r += 1

    tbl2 = wsj.tables["Tabela26"]
    tbl2.ref = f"A1:AH{n_jdd_rows + 1}"
    if tbl2.autoFilter is not None:
        tbl2.autoFilter.ref = tbl2.ref

    # ---------------- MERCADOS ----------------
    wsm = wb["Mercados"]
    mer_tpl = _capture_template(wsm, 2, MER_FORMULA_COLS)
    old_last = int(re.search(r"\d+", wsm.tables["Tabela37"].ref.split(":")[1]).group())
    if old_last > n_mer_rows + 1:
        wsm.delete_rows(n_mer_rows + 2, old_last - (n_mer_rows + 1))
    elif old_last < n_mer_rows + 1:
        wsm.insert_rows(old_last + 1, (n_mer_rows + 1) - old_last)

    r = 2
    for g in games:
        for row in g["mercados_rows"]:
            jogo, mercado, tipo, c1, c2, pfont, odd = row
            wsm.cell(r, 1, jogo)
            wsm.cell(r, 2, mercado)
            wsm.cell(r, 3, tipo)
            wsm.cell(r, 4, c1)
            wsm.cell(r, 5, c2)
            wsm.cell(r, 7, pfont)
            wsm.cell(r, 11, odd)
            for col, (txt, is_arr) in mer_tpl.items():
                new_txt = clone_formula(txt, 2, r)
                if col in (20, 21):  # T, U -> ranges de Jogos do Dia
                    new_txt = fix_range_end(new_txt, 5, n_jdd_rows + 1)
                _set_formula(wsm, r, col, new_txt, is_arr)
            r += 1

    tbl3 = wsm.tables["Tabela37"]
    tbl3.ref = f"A1:Q{n_mer_rows + 1}"
    if tbl3.autoFilter is not None:
        tbl3.autoFilter.ref = tbl3.ref

    # ---------------- FONTES ----------------
    wsf = wb["Fontes"]
    max_existing = wsf.max_row
    for rr in range(2, max_existing + 1):
        for c in range(1, 9):
            wsf.cell(rr, c).value = None
    for i, g in enumerate(games):
        rr = 2 + i
        f = g["fontes"]
        jogo = f"{g['teamA']} x {g['teamB']}"
        wsf.cell(rr, 1, jogo)
        wsf.cell(rr, 2, f["placar_modal"])
        wsf.cell(rr, 3, f["p_a"])
        wsf.cell(rr, 4, f["p_empate"])
        wsf.cell(rr, 5, f["p_b"])
        wsf.cell(rr, 6, f["consenso"])
        wsf.cell(rr, 7, f["melhor_aposta"])
        wsf.cell(rr, 8, f["atencao"])

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 4. AJUSTE DE MANDO NO λ (Poisson) — mecânica genérica (versão B com shrinkage)
# ---------------------------------------------------------------------------
# PROBLEMA: o λ (Gols Pró/Contra na aba 'Jogos do Dia') é a média ponderada dos
# gols do time no histórico, misturando jogos em CASA e FORA. Numa liga de
# pontos corridos isso subestima o mandante e superestima o visitante,
# distorcendo TODOS os mercados de gols (Under/Over, BTTS, AH, quem marca).
#
# SOLUÇÃO (shrinkage): ponderar o histórico de cada time pelo MANDO. Os jogos
# no mando que o time terá HOJE pesam 1; os do mando oposto pesam k.
#   k = 1.0  -> comportamento antigo (sem mando)
#   k = 0.0  -> split casa/fora puro (OVERFIT: amostra pequena por mando)
#   k = 0.35 -> valor usado como ponto de partida na Série A/B (ver ligas.py,
#               mando_shrink_k_default) -- calibrar por liga, não copiar cego.
#
# IMPLEMENTAÇÃO SEGURA: NÃO altera a fórmula do λ nem o POISSON. A fórmula do λ
# já multiplica por Times!AK (Peso Final = AG*AH*AI). Este pós-processador:
#   1. escreve o mando do jogo (Casa/Fora por time) numa coluna nova de 'Jogos
#      do Dia' (AK, "Mando Jogo");
#   2. cria em Times a coluna AL ("Fator Mando") = 1 se o mando da linha bate
#      com o mando do time no jogo de hoje, senão k;
#   3. reescreve Times!AK para AG*AH*AI*AL.
# Só faz sentido em competições com mando real -- não chamar para torneios
# neutros (ver ligas.LigaConfig.neutro / mando_shrink_k_default).

def aplicar_ajuste_mando(xlsx_path, k, home_teams):
    """Aplica o ajuste de mando no λ da planilha JÁ gerada por build_workbook.

    xlsx_path : caminho do .xlsx de saída (será reescrito no lugar).
    k         : fator de encolhimento dos jogos do mando oposto (0..1). Pegue
                o default recomendado da liga com
                ligas.mando_shrink_k_default(liga) em vez de chutar um valor.
    home_teams: lista/set com os nomes dos MANDANTES (time que joga em casa) de
                cada jogo do dia. Necessário para saber, por time, se hoje ele é
                Casa ou Fora. Ex.: {"América Mineiro", "Ceará"}.

    Rodar SEMPRE depois de build_workbook e ANTES do recalc.py. NÃO chamar
    para torneios neutros (Copa) -- lá não há mandante real.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    if not home_teams:
        raise ValueError("Informe home_teams (mandantes do dia) para o ajuste de mando.")
    home_teams = set(home_teams)

    wb = load_workbook(xlsx_path)
    jd = wb["Jogos do Dia"]
    tm = wb["Times"]

    # (1) coluna "Mando Jogo" em Jogos do Dia (col AK = 37), por time
    jd_mando_col = 37  # AK, primeira livre após AJ(Chave)=36
    jd.cell(1, jd_mando_col).value = "Mando Jogo"
    last_jd = jd.max_row
    for r in range(2, last_jd + 1):
        time = jd.cell(r, 2).value  # col B = Time
        if not time:
            continue
        jd.cell(r, jd_mando_col).value = "Casa" if time in home_teams else "Fora"

    # (2) coluna "Fator Mando" em Times (col AL = 38)
    #     AL = 1 se Mando da linha (E) == Mando do time no jogo de hoje, senão k.
    #     O mando do time hoje é lido de 'Jogos do Dia' via PROCV pelo nome.
    tm_fator_col = 38  # AL
    tm.cell(1, tm_fator_col).value = "Fator Mando"
    jd_key_range = f"'Jogos do Dia'!$B$2:$B${last_jd}"
    jd_mando_range = f"'Jogos do Dia'!${get_column_letter(jd_mando_col)}$2:${get_column_letter(jd_mando_col)}${last_jd}"
    last_tm = tm.max_row
    for r in range(2, last_tm + 1):
        nome = tm.cell(r, 1).value  # A = Time
        if nome in (None, ""):
            continue
        # INDEX/MATCH (compatível com LibreOffice, sem XLOOKUP)
        lookup = f"INDEX({jd_mando_range},MATCH($A{r},{jd_key_range},0))"
        tm.cell(r, tm_fator_col).value = f"=IFERROR(IF($E{r}={lookup},1,{k}),1)"

    # (3) Peso Final AK = AG*AH*AI*AL
    for r in range(2, last_tm + 1):
        if tm.cell(r, 1).value in (None, ""):
            continue
        tm.cell(r, 37).value = f"=AG{r}*AH{r}*AI{r}*AL{r}"

    # registrar o parâmetro k na aba Parâmetros (rastreabilidade)
    if "Parâmetros" in wb.sheetnames:
        pa = wb["Parâmetros"]
        row = pa.max_row + 2
        pa.cell(row, 1).value = "Fator mando (k)"
        pa.cell(row, 2).value = k
        pa.cell(row, 3).value = "Encolhimento do mando oposto no λ (0=split puro, 1=sem mando)."

    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------------------------------------------------------
# 5. FÓRMULA "PRÓ/CONTRA" ROBUSTA — MODELO 16/07/2026 (v2, sem LET/array)
# ---------------------------------------------------------------------------
# Preserva as 3 melhorias estatísticas do modelo de Lucas (média ponderada,
# desvio-padrão com correção de reliability weights, corte parametrizado
# unilateral/bilateral), mas SEM LET e SEM ArrayFormula — a versão original
# do Lucas usava LET com variáveis internas array (m, x, w, sel), o que exige
# CSE (Ctrl+Shift+Enter) pra não sofrer injeção de '@' pelo Excel ao reabrir.
# Só que CSE dentro de uma Tabela do Excel (Tabela26/Tabela37, ListObject) é
# gatilho conhecido de "Encontramos um problema em um conteúdo..." — bug
# real, diagnosticado e corrigido na Série A.
#
# Solução: expandir a lógica em DUAS colunas auxiliares por estatística, cada
# uma um SUMPRODUCT simples (nunca precisa de LET nem array, porque toda
# variável "array" fica sempre dentro de um SUMPRODUCT, nunca vira uma
# referência nomeada solta que o Excel possa reinterpretar):
#   aux μ (col. BC, BE, BG... pares)  = SUMPRODUCT(válido*valor*peso) / SUMPRODUCT(válido*peso)
#   aux σ (col. BD, BF, BH... ímpares) = SQRT(SUMPRODUCT(válido*peso*(valor-μ)²) / (ΣW-ΣW²/ΣW))
# O indicador principal só referencia as duas células auxiliares (μ, σ) e o
# corte parametrizado (Parâmetros!B10/B11) — nunca recalcula nada por conta
# própria, então também é SUMPRODUCT puro, sem LET.
#
# Parâmetros de corte (Parâmetros!B10/B11, mesmos valores/semântica do
# modelo do Lucas):
#   B10 "Limite para corte unilateral (média)" (padrão 4): se μ <= B10, corte
#        UNILATERAL (só remove valores ACIMA de μ+limite — não penaliza jogos
#        de placar baixo, comum em gols/cartões/chutes-a-gol; só poda picos).
#   B11 "Multiplicador do desvio-padrão" (padrão 2.5): limite = B11*σ. Acima
#        da média-piso (B10), corte BILATERAL (ABS(x-μ)<=limite).
#
# Rodar SEMPRE por último (build_workbook -> aplicar_ajuste_mando ->
# montar_aba_padroes -> recalc.py -> aplicar_formula_pro_contra ->
# sanitizar_v_vazio -> present_files).

_PRO_CONTRA_MAP = {
    "Gols Pró": "F", "Gols Contra": "G",
    "Cartões Pró": "L", "Cartões Contra": "M",
    "Escanteios Pró": "J", "Escanteios Contra": "K",
    "Chutes Pró": "N", "Chutes Contra": "O",
    "Chutes Gol Pró": "P", "Chutes Gol Contra": "Q",
    "Gols 1T Pró": "U", "Gols 1T Contra": "V",
}

# Duas auxiliares por estatística: MU (média ponderada) e SD (desvio-padrão
# ponderado com correção de reliability weights). BC em diante, sequencial.
AUX_START_COL = 55  # BC

CORTE_LIMITE_UNILATERAL_DEFAULT = 4     # Parâmetros!B10
CORTE_MULTIPLICADOR_DP_DEFAULT = 2.5    # Parâmetros!B11


def _valid_expr(row, last):
    A = f"Times!$A$2:$A${last}"
    AG = f"Times!$AG$2:$AG${last}"
    AH = f"Times!$AH$2:$AH${last}"
    return f"({A}=$B{row})*({AG}>=65%)*({AH}>=65%)"


def _formula_mu(row, src_col, last):
    """Média PONDERADA (pelo Peso Final Times!AK) — coluna auxiliar MU.
    SUMPRODUCT puro, sem LET, sem array."""
    AK = f"Times!$AK$2:$AK${last}"
    SRC = f"Times!${src_col}$2:${src_col}${last}"
    valid = _valid_expr(row, last)
    return f"=IFERROR(SUMPRODUCT({valid}*{SRC}*{AK})/SUMPRODUCT({valid}*{AK}),0)"


def _formula_sd(row, src_col, last, mu_ref):
    """Desvio-padrão PONDERADO com correção de reliability weights — coluna
    auxiliar SD. Referencia a célula MU (mu_ref, ex. '$BC2'). SUMPRODUCT
    puro, sem LET, sem array.
    sd = SQRT( SOMA(peso*(x-mu)^2) / (SOMA(peso) - SOMA(peso^2)/SOMA(peso)) )"""
    AK = f"Times!$AK$2:$AK${last}"
    SRC = f"Times!${src_col}$2:${src_col}${last}"
    valid = _valid_expr(row, last)
    somaW = f"SUMPRODUCT({valid}*{AK})"
    somaWq = f"SUMPRODUCT({valid}*{AK}*{AK})"
    denom = f"(({somaW})-({somaWq})/({somaW}))"
    return (
        f"=IFERROR(IF({denom}>0,"
        f"SQRT(SUMPRODUCT({valid}*{AK}*({SRC}-{mu_ref})^2)/{denom}),0),0)"
    )


def _formula_pro_contra(row, src_col, last, mu_ref, sd_ref):
    """Indicador principal: corte unilateral/bilateral parametrizado sobre a
    média ponderada, com desvio-padrão ponderado. SUMPRODUCT puro, sem LET,
    sem array — mu_ref/sd_ref são referências às células auxiliares (ex.
    '$BC2', '$BD2').

    CRÍTICO (bug real, achado por Lucas): a máscara de seleção NÃO pode usar
    SE()/IF() com a range de dados (SRC) dentro dos dois ramos — quando o
    Excel reabre um arquivo escrito por fora dele, ele injeta '@' (interseção
    implícita) na referência de range que está dentro de um IF/SE aninhado,
    mesmo sem LET e sem array formula. Isso silenciosamente troca a range
    inteira por só a célula da mesma linha, zerando o resultado pra times com
    poucos jogos válidos (n pequeno). A correção é eliminar o SE() da máscara
    e usar combinação ARITMÉTICA (multiplicação por 0/1), que o Excel nunca
    reinterpreta com '@' porque não há mais nenhum IF/SE envolvendo a range
    SRC — SRC só aparece direto dentro de SUMPRODUCT, nunca dentro de um IF
    aninhado."""
    AK = f"Times!$AK$2:$AK${last}"
    SRC = f"Times!${src_col}$2:${src_col}${last}"
    valid = _valid_expr(row, last)
    n = f"SUMPRODUCT({valid})"
    lim = f"Parâmetros!$B$11*{sd_ref}"
    uni = f"({mu_ref}<=Parâmetros!$B$10)"  # escalar: TRUE/FALSE -> 1/0
    sel_unilateral = f"({SRC}<={mu_ref}+{lim})"
    sel_bilateral = f"(ABS({SRC}-{mu_ref})<={lim})"
    # combinação aritmética em vez de SE()/IF() -- nunca envolve SRC num IF
    sel = f"({uni}*{sel_unilateral}+(1-{uni})*{sel_bilateral})"
    return (
        f"=IFERROR(IF({n}<3,"
        f"SUMPRODUCT({valid}*{SRC}*{AK})/SUMPRODUCT({valid}*{AK}),"
        f"SUMPRODUCT({valid}*{sel}*{SRC}*{AK})/SUMPRODUCT({valid}*{sel}*{AK})),0)"
    )


def _garantir_parametros_corte(wb, limite_unilateral=CORTE_LIMITE_UNILATERAL_DEFAULT,
                                multiplicador_dp=CORTE_MULTIPLICADOR_DP_DEFAULT):
    """Garante Parâmetros!B10/B11 (corte unilateral/multiplicador de DP) nas
    posições exatas referenciadas pela fórmula do indicador. Idempotente."""
    if "Parâmetros" not in wb.sheetnames:
        raise ValueError("Aba 'Parâmetros' não encontrada — necessária para B10/B11.")
    pa = wb["Parâmetros"]
    pa.cell(10, 1).value = "Limite para corte unilateral (média)"
    pa.cell(10, 2).value = limite_unilateral
    pa.cell(10, 3).value = "Até esta média, o filtro remove somente valores acima do limite superior."
    pa.cell(11, 1).value = "Multiplicador do desvio-padrão"
    pa.cell(11, 2).value = multiplicador_dp
    pa.cell(11, 3).value = (f"Limite estatístico de {multiplicador_dp}x DP; acima da média "
                             f"{limite_unilateral}, o corte permanece bilateral.")


def aplicar_formula_pro_contra(xlsx_path, limite_unilateral=CORTE_LIMITE_UNILATERAL_DEFAULT,
                                multiplicador_dp=CORTE_MULTIPLICADOR_DP_DEFAULT):
    """Reescreve as 12 colunas Pró/Contra da aba 'Jogos do Dia' com o modelo
    16/07/2026 (média ponderada + DP ponderado com reliability weights +
    corte unilateral/bilateral parametrizado) — SEM LET, SEM ArrayFormula,
    100% SUMPRODUCT clássico. Nunca mais fórmula array dentro de Tabela.

    Faz, nesta ordem:
      1. Garante Parâmetros!B10/B11 (não duplica).
      2. Escreve as 24 colunas auxiliares MU/SD (BC:BZ) — 2 por estatística.
      3. Escreve os 12 indicadores, cada um referenciando suas 2 auxiliares.

    🚨 NUNCA rodar recalc.py (ou qualquer coisa que abra o arquivo no
    LibreOffice) DEPOIS desta função. Ordem definitiva: build_workbook →
    aplicar_ajuste_mando → montar_aba_padroes → recalc.py →
    aplicar_formula_pro_contra → sanitizar_v_vazio → present_files."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(xlsx_path)
    jd = wb["Jogos do Dia"]
    tm = wb["Times"]

    _garantir_parametros_corte(wb, limite_unilateral, multiplicador_dp)

    last = 1
    for r in range(2, tm.max_row + 1):
        if tm.cell(r, 1).value not in (None, ""):
            last = r

    hdr = {jd.cell(1, c).value: c for c in range(1, jd.max_column + 1)}
    last_jd = 1
    for r in range(2, jd.max_row + 1):
        if jd.cell(r, 2).value not in (None, ""):
            last_jd = r

    for i, (nome, src) in enumerate(_PRO_CONTRA_MAP.items()):
        if nome not in hdr:
            continue
        col = hdr[nome]
        mu_col = AUX_START_COL + i * 2
        sd_col = mu_col + 1
        mu_letter = get_column_letter(mu_col)
        sd_letter = get_column_letter(sd_col)
        jd.cell(1, mu_col).value = f"_mu {nome}"
        jd.cell(1, sd_col).value = f"_sd {nome}"
        for r in range(2, last_jd + 1):
            if jd.cell(r, 2).value in (None, ""):
                continue
            mu_ref = f"${mu_letter}{r}"
            sd_ref = f"${sd_letter}{r}"
            jd.cell(r, mu_col).value = _formula_mu(r, src, last)
            jd.cell(r, sd_col).value = _formula_sd(r, src, last, mu_ref)
            jd.cell(r, col).value = _formula_pro_contra(r, src, last, mu_ref, sd_ref)

    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------------------------------------------------------
# 6. P.FONTE ESTIMADO — número (%) + destaque de cor, nunca texto na célula
# ---------------------------------------------------------------------------
# Escrever "0.35 [ESTIMADO]" na célula de P.Font transforma a célula em TEXTO
# e quebra toda fórmula que soma/multiplica essa coluna (Edge, P.Comb etc.).
# P.Font e Odd SEMPRE como número puro (float). Quando o valor é estimado
# (sem fonte externa real), sinalizar com PREENCHIMENTO DE COR na célula.

ESTIMADO_FILL_COLOR = "FFF2CC"  # amarelo claro


def marcar_estimados(xlsx_path, game_index, linhas_estimadas, sheet="Mercados",
                      cols=(7, 11)):
    """Aplica destaque de cor nas células de P.Font (col 7) e Odd (col 11) do
    Mercados para linhas cujo valor foi ESTIMADO (sem fonte externa real).

    game_index      : índice do jogo no dia (0-based).
    linhas_estimadas: posições (1-20) dentro do bloco de 20 mercados daquele
                       jogo que são estimadas."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    wb = load_workbook(xlsx_path)
    ws = wb[sheet]
    fill = PatternFill(start_color=ESTIMADO_FILL_COLOR, end_color=ESTIMADO_FILL_COLOR,
                        fill_type="solid")
    base_row = game_index * 20 + 1
    for pos in linhas_estimadas:
        r = base_row + pos
        for c in cols:
            ws.cell(r, c).fill = fill
    wb.save(xlsx_path)
    return xlsx_path


def sanitizar_v_vazio(xlsx_path):
    """Remove tags <v></v> vazias (sem valor em cache) de todas as abas do
    arquivo. openpyxl às vezes deixa essas tags vazias em células de fórmula
    quando o workbook é reaberto/resalvo várias vezes — isso é XML inválido
    pro Excel: dispara 'Encontramos um problema em um conteúdo...' e o
    reparo automático às vezes APAGA a fórmula. Fórmula sem cache é válida
    SEM a tag <v> — o Excel recalcula normalmente ao abrir.

    🚨 ÚLTIMO passo do fluxo, depois de aplicar_formula_pro_contra, antes de
    present_files. Não usa openpyxl (edita o XML do zip direto)."""
    import zipfile, os
    tmp = xlsx_path + ".tmp"
    total = 0
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                n = text.count("<v></v>")
                if n:
                    total += n
                    text = text.replace("<v></v>", "")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, xlsx_path)
    return total
