"""
Converte AVULSO (não mexe na skill) as 12 fórmulas Pró/Contra da SerieB_13jul
de LET/FILTER (array dinâmico, quebra no Excel mobile) para SUMPRODUCT + coluna
auxiliar de média. SUMPRODUCT não é array dinâmico: sobrevive ao openpyxl e abre
em qualquer Excel.

Lógica preservada (idêntica ao LET do Lucas):
  valid = (Times.A==time) * (Times.AG>=65%) * (Times.AH>=65%)
  avg   = SUMPRODUCT(valid*val) / SUMPRODUCT(valid)          [coluna auxiliar]
  sd    = SQRT( SUMPRODUCT(valid*(val-avg)^2) / (SUMPRODUCT(valid)-1) )
  mask  = |val-avg| <= sd        (com n>=3; n<3 usa todos)
  result= SUMPRODUCT(valid*mask*val*peso) / SUMPRODUCT(valid*mask*peso)
onde peso = Times.AK (já inclui o fator de mando AL).
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH='/mnt/user-data/outputs/SerieB_13jul.xlsx'
wb=load_workbook(PATH)
jd=wb['Jogos do Dia']; tm=wb['Times']

# última linha de dados em Times
last=1
for r in range(2,tm.max_row+1):
    if tm.cell(r,1).value not in (None,''): last=r

# mercado -> (col destino em JD, col fonte em Times)
MAPA={'Gols Pró':(15,'F'),'Gols Contra':(16,'G'),'Cartões Pró':(18,'L'),'Cartões Contra':(19,'M'),
'Escanteios Pró':(21,'J'),'Escanteios Contra':(22,'K'),'Chutes Pró':(24,'N'),'Chutes Contra':(25,'O'),
'Chutes Gol Pró':(27,'P'),'Chutes Gol Contra':(28,'Q'),'Gols 1T Pró':(30,'U'),'Gols 1T Contra':(31,'V')}

# linhas de jogo (com time preenchido)
game_rows=[r for r in range(2,jd.max_row+1) if jd.cell(r,2).value not in (None,'')]

AUX_START=55  # coluna BC em diante, uma por mercado (guarda o avg)
A=f"Times!$A$2:$A${last}"; AG=f"Times!$AG$2:$AG${last}"; AH=f"Times!$AH$2:$AH${last}"; AK=f"Times!$AK$2:$AK${last}"

for idx,(nome,(dcol,src)) in enumerate(MAPA.items()):
    auxcol=AUX_START+idx
    auxL=get_column_letter(auxcol)
    SRC=f"Times!${src}$2:${src}${last}"
    # cabeçalho aux
    jd.cell(1,auxcol).value=f"_avg {nome}"
    for r in game_rows:
        tmn=f"$B{r}"
        valid=f"(({A}={tmn})*({AG}>=65%)*({AH}>=65%))"
        # avg na coluna aux
        avg_formula=f"=IFERROR(SUMPRODUCT({valid}*{SRC})/SUMPRODUCT({valid}),0)"
        jd.cell(r,auxcol).value=avg_formula
        avg_ref=f"${auxL}{r}"
        # n e sd inline
        n=f"SUMPRODUCT({valid})"
        var=f"SUMPRODUCT({valid}*({SRC}-{avg_ref})^2)/MAX({n}-1,1)"
        sd=f"SQRT({var})"
        # mask: se n<3 usa todos; senão |val-avg|<=sd
        # (val-avg) e comparação viram parte do SUMPRODUCT como (ABS(SRC-avg)<=sd)
        maskcond=f"(ABS({SRC}-{avg_ref})<={sd})"
        # se n<3, mask=1 para todos os válidos -> uso IF fora
        num=f"SUMPRODUCT({valid}*IF({n}<3,1,{maskcond})*{SRC}*{AK})"
        den=f"SUMPRODUCT({valid}*IF({n}<3,1,{maskcond})*{AK})"
        # IF dentro de SUMPRODUCT força array -> evitar: separar em dois SUMPRODUCT via SE(n<3;...;...)
        full=(f"=IFERROR(IF({n}<3,"
              f"SUMPRODUCT({valid}*{SRC}*{AK})/SUMPRODUCT({valid}*{AK}),"
              f"SUMPRODUCT({valid}*{maskcond}*{SRC}*{AK})/SUMPRODUCT({valid}*{maskcond}*{AK})),0)")
        jd.cell(r,dcol).value=full

wb.save(PATH)
print("convertido para SUMPRODUCT")
