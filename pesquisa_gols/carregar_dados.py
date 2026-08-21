"""
Lê as abas Jogos, Snapshots, Stats_Finais e Matriz da planilha original e
devolve estruturas Python simples (listas/dicionários), sem depender de
pandas — o app principal (ligas_live_app) também não usa, e para o volume
de dados aqui (240 jogos, ~1.700 snapshots) não faz falta.
"""
import re

import openpyxl

import config
from matriz_padrao import CORRELACAO_GOLS_PADRAO


def _linhas(aba):
    """Gera dicionários {cabeçalho: valor} para cada linha de dados de uma aba."""
    linhas = aba.iter_rows(values_only=True)
    cabecalho = next(linhas)
    for linha in linhas:
        if linha[0] is None:
            continue
        yield dict(zip(cabecalho, linha))


def carregar_jogos(wb):
    """fixture_id -> {rodada, data_hora, time_casa, time_fora, finalizado}"""
    jogos = {}
    for row in _linhas(wb["Jogos"]):
        jogos[int(row["fixture_id"])] = {
            "rodada": row["rodada"],
            "data_hora": row["data_hora"],
            "time_casa": row["time_casa"],
            "time_fora": row["time_fora"],
            "finalizado": bool(row["finalizado"]),
        }
    return jogos


def carregar_gols_finais(wb):
    """
    fixture_id -> total de gols na partida (goals_casa + goals_fora).

    Alguns jogos têm goals_casa/goals_fora vazios em Stats_Finais (falha da
    fonte) mesmo já finalizados — nesses casos usa o placar do snapshot 'FT'
    em Snapshots como fallback, que está sempre preenchido.
    """
    gols_finais = {}
    for row in _linhas(wb["Stats_Finais"]):
        casa = row.get("goals_casa")
        fora = row.get("goals_fora")
        if casa is None or fora is None:
            continue
        gols_finais[int(row["fixture_id"])] = int(casa) + int(fora)

    for row in _linhas(wb["Snapshots"]):
        if row.get("marco_min") != "FT":
            continue
        fixture_id = int(row["fixture_id"])
        if fixture_id in gols_finais:
            continue
        placar_casa = row.get("placar_casa")
        placar_fora = row.get("placar_fora")
        if placar_casa is None or placar_fora is None:
            continue
        gols_finais[fixture_id] = int(placar_casa) + int(placar_fora)

    return gols_finais


def _normalizar_indicador(texto):
    """'shots_total / goal_attempts' -> ['shots_total', 'goal_attempts']; remove anotações tipo '(trend)'."""
    partes = texto.split("/")
    nomes = []
    for parte in partes:
        parte = re.sub(r"\([^)]*\)", "", parte).strip()
        if parte:
            nomes.append(parte)
    return nomes


def carregar_matriz(wb):
    """
    nome_base_do_indicador (sem _casa/_fora) -> correlação com Gols ('Direto', 'Indireto forte', ...).

    A aba Matriz é opcional: se o arquivo de entrada não tiver uma (caso normal
    para uma exportação de temporada nova — Matriz é conhecimento fixo sobre os
    indicadores, não dado da temporada), usa o padrão embutido em matriz_padrao.py.
    """
    if "Matriz" not in wb.sheetnames:
        return dict(CORRELACAO_GOLS_PADRAO)

    aba = wb["Matriz"]
    todas_linhas = list(aba.iter_rows(values_only=True))
    # a aba tem título/subtítulo antes do cabeçalho de verdade — acha a linha
    # que começa com "Indicador" em vez de assumir que é a primeira.
    idx_cabecalho = next(
        i for i, linha in enumerate(todas_linhas) if linha and linha[0] == "Indicador"
    )
    cabecalho = todas_linhas[idx_cabecalho]
    correlacoes = {}
    for linha in todas_linhas[idx_cabecalho + 1:]:
        if not linha or linha[0] is None:
            continue
        row = dict(zip(cabecalho, linha))
        indicador = row.get("Indicador")
        correlacao_gols = row.get("Gols")
        if not indicador or not correlacao_gols:
            continue
        for nome in _normalizar_indicador(indicador):
            correlacoes[nome] = correlacao_gols
    return correlacoes


def bases_disponiveis(wb):
    """Nomes-base de colunas da aba Snapshots (sem sufixo _casa/_fora)."""
    cabecalho = next(wb["Snapshots"].iter_rows(values_only=True))
    bases = set()
    for nome_coluna in cabecalho:
        if not isinstance(nome_coluna, str):
            continue
        for sufixo in ("_casa", "_fora"):
            if nome_coluna.endswith(sufixo):
                bases.add(nome_coluna[: -len(sufixo)])
    return bases


def estatisticas_candidatas(bases, matriz):
    """
    Nomes-base de estatísticas (interseção entre os arquivos, quando há mais de
    um) que têm correlação registrada com Gols na Matriz e não são o próprio
    placar.
    """
    return sorted(
        b for b in bases
        if b in matriz and b not in config.INDICADORES_EXCLUIDOS
    )


def carregar_snapshots(wb, candidatas):
    """
    Lista de snapshots: um dict por (fixture_id, minuto), com o placar somado
    (gols_momento), o minuto e o valor casa+fora de cada estatística candidata.
    """
    snapshots = []
    for row in _linhas(wb["Snapshots"]):
        marco_min = row.get("marco_min")
        if not isinstance(marco_min, (int, float)):
            # 'FT' (fim de jogo) não é um minuto fixo comparável entre partidas
            # — a pesquisa trabalha só com os marcos ao vivo (15, 30, 45, 60, 75, 90).
            continue
        placar_casa = row.get("placar_casa") or 0
        placar_fora = row.get("placar_fora") or 0
        snap = {
            "fixture_id": int(row["fixture_id"]),
            "minuto": int(marco_min),
            "gols_momento": int(placar_casa) + int(placar_fora),
        }
        for stat in candidatas:
            casa = row.get(f"{stat}_casa") or 0
            fora = row.get(f"{stat}_fora") or 0
            try:
                snap[stat] = float(casa) + float(fora)
            except (TypeError, ValueError):
                snap[stat] = 0.0
        snapshots.append(snap)
    return snapshots


def carregar_tudo(caminhos=None):
    """
    Carrega tudo que buscar_condicoes.py precisa, a partir de um arquivo ou uma
    lista de arquivos (config.ARQUIVOS_ENTRADA) — permite juntar várias
    temporadas/ligas num único dataset, desde que fixture_id não se repita
    entre elas (times/temporadas diferentes normalmente garantem isso).
    """
    if caminhos is None:
        caminhos = config.ARQUIVOS_ENTRADA
    elif isinstance(caminhos, str):
        caminhos = [caminhos]

    workbooks = [openpyxl.load_workbook(c, read_only=True, data_only=True) for c in caminhos]

    matriz = {}
    for wb in workbooks:
        matriz.update(carregar_matriz(wb))

    bases_comuns = None
    for wb in workbooks:
        bases = bases_disponiveis(wb)
        bases_comuns = bases if bases_comuns is None else (bases_comuns & bases)
    candidatas = estatisticas_candidatas(bases_comuns, matriz)

    jogos, gols_finais, snapshots = {}, {}, []
    for caminho, wb in zip(caminhos, workbooks):
        jogos_arquivo = carregar_jogos(wb)
        repetidos = set(jogos_arquivo) & set(jogos)
        if repetidos:
            raise ValueError(
                f"{caminho}: {len(repetidos)} fixture_id já presentes em outro arquivo carregado "
                f"(ex.: {sorted(repetidos)[:3]}) — não dá para juntar sem sobrescrever jogos."
            )
        jogos.update(jogos_arquivo)
        gols_finais.update(carregar_gols_finais(wb))
        snapshots.extend(carregar_snapshots(wb, candidatas))

    return {
        "jogos": jogos,
        "gols_finais": gols_finais,
        "matriz": matriz,
        "candidatas": candidatas,
        "snapshots": snapshots,
    }


if __name__ == "__main__":
    dados = carregar_tudo()
    print(f"{len(dados['jogos'])} jogos, {len(dados['snapshots'])} snapshots, "
          f"{len(dados['candidatas'])} estatísticas candidatas (com correlação registrada na Matriz)")
    print("Candidatas:", ", ".join(dados["candidatas"]))
