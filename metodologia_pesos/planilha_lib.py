"""
planilha_lib.py — biblioteca reutilizável para montar Copa_Xjul.xlsx
(skill copa-planilha-dia)

Uso típico (ver SKILL.md para o passo a passo completo):

    import sys
    sys.path.insert(0, "/mnt/skills/user/copa-planilha-dia/scripts")
    from planilha_lib import *

    df = load_all_matches()  # csvs em /mnt/project/*.csv
    hist_A = get_historico("France", df)
    hist_B = get_historico("Morocco", df)
    attach_estilo(hist_A); attach_estilo(hist_B)   # usa data/estilos_selecoes.json
    # ... preencher estilos novos manualmente no dict antes de attach_estilo se faltar

    games = [
        dict(teamA="France", teamB="Morocco", hist_A=hist_A, hist_B=hist_B,
             jogo_label="France x Morocco", data_jogo="09/07/2026",
             jdd_A=dict(...), jdd_B=dict(...), mercados=[...], fontes=dict(...))
    ]
    build_workbook(games, template_path=".../Copa_Template_Simplificado.xlsx",
                   output_path="/mnt/user-data/outputs/Copa_9jul.xlsx")

Este módulo cuida da parte MECÂNICA (dados históricos dos CSVs, clonagem de
fórmulas, redimensionamento das 3 tabelas do Excel). A parte de JULGAMENTO
(quais jogos, odds do dia, estilo, probabilidades das fontes) continua sendo
feita pelo Claude a cada sessão — ver SKILL.md.
"""
import re
import json
import glob
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

CSV_GLOB_DEFAULT = "/mnt/project/*.csv"
ESTILO_DB_PATH = "/mnt/skills/user/copa-planilha-dia/data/estilos_selecoes.json"

# Alias de nomes divergentes entre CSV do FootyStats e nomenclatura usada nas
# planilhas / casas de aposta. Adicionar aqui sempre que aparecer um novo caso
# (ver seção "Nomenclaturas CSV a verificar" no briefing_copa.md).
CSV_ALIAS = {
    "USMNT": "USA",       # CSV usa "USA", planilha usa "USMNT"
    "Congo DR": "DR Congo",
}


# ---------------------------------------------------------------------------
# 1. DADOS HISTÓRICOS (Times sheet)
# ---------------------------------------------------------------------------

def load_all_matches(csv_glob=CSV_GLOB_DEFAULT):
    """Carrega e concatena todos os CSVs do FootyStats em /mnt/project/."""
    dfs = []
    for f in glob.glob(csv_glob):
        try:
            d = pd.read_csv(f)
            d["__src"] = f.split("/")[-1]
            dfs.append(d)
        except Exception as e:
            print("ERRO ao ler", f, e)
    df = pd.concat(dfs, ignore_index=True, sort=False)
    return df[df["status"] == "complete"]


def _comp_name(src):
    """Mapeia nome de arquivo CSV -> (nome da competição, mando forçado ou None).
    Mando forçado 'Neutro' para Copa do Mundo e Eurocopa (jogos em local
    neutro na convenção usada nesta planilha)."""
    s = src.lower()
    if "worldcup" in s:
        return ("Copa do Mundo 2026", "Neutro")
    if "eurochampionship" in s:
        return ("Eurocopa 2024", "Neutro")
    if "nationsleague" in s:
        return ("Nations League", None)
    if "friendliesmatches2026" in s:
        return ("Amistoso/Preparação 2026", None)
    if "friendliesmatches2025" in s:
        return ("Amistoso Internacional", None)
    if "qualificationeurope" in s:
        return ("Elim. Copa (Europa)", None)
    if "qualificationafrica" in s:
        return ("Elim. Copa (África)", None)
    if "qualificationconcacaf" in s:
        return ("Elim. Copa (CONCACAF)", None)
    if "qualificationasia" in s:
        return ("Elim. Copa (Ásia)", None)
    if "qualificationsouthamerica" in s:
        return ("Elim. Copa (Am. do Sul)", None)
    if "qualificationintercontinental" in s:
        return ("Repescagem Intercontinental", "Neutro")
    return (src, None)


def get_historico(team, df, n=15, alias=None):
    """Retorna os últimos n jogos completos de `team`, mais recentes primeiro,
    já na perspectiva do time (Pró/Contra, mando, favoritismo normalizado).

    `alias`: nome alternativo a procurar no CSV se `team` não bater
    (ex.: get_historico("USMNT", df, alias="USA")).
    Sinaliza (print) se encontrar menos que n jogos — nesse caso, reduzir
    stake/qualificar a análise conforme Regra 2 do protocolo.
    """
    lookup = alias or team
    sub = df[(df["home_team_name"] == lookup) | (df["away_team_name"] == lookup)].copy()
    sub = sub.sort_values("timestamp", ascending=False).head(n)
    if len(sub) < n:
        print(f"⚠️  {team}: apenas {len(sub)}/{n} jogos encontrados no CSV — "
              f"considerar amostra parcial (Regra 2 do protocolo).")
    out = []
    for _, r in sub.iterrows():
        is_home = r["home_team_name"] == lookup
        opp = r["away_team_name"] if is_home else r["home_team_name"]
        d = datetime.datetime.strptime(r["date_GMT"].split(" - ")[0], "%b %d %Y")
        comp, forced_mando = _comp_name(r["__src"])
        mando = forced_mando if forced_mando else ("Casa" if is_home else "Fora")

        def pick(a, b):
            return a if is_home else b

        gf, ga = pick(r["home_team_goal_count"], r["away_team_goal_count"]), pick(r["away_team_goal_count"], r["home_team_goal_count"])
        xgf, xga = pick(r["team_a_xg"], r["team_b_xg"]), pick(r["team_b_xg"], r["team_a_xg"])
        cf, ca = pick(r["home_team_corner_count"], r["away_team_corner_count"]), pick(r["away_team_corner_count"], r["home_team_corner_count"])
        yf = pick(r["home_team_yellow_cards"] + r["home_team_red_cards"], r["away_team_yellow_cards"] + r["away_team_red_cards"])
        ya = pick(r["away_team_yellow_cards"] + r["away_team_red_cards"], r["home_team_yellow_cards"] + r["home_team_red_cards"])
        sf, sa = pick(r["home_team_shots"], r["away_team_shots"]), pick(r["away_team_shots"], r["home_team_shots"])
        sotf, sota = pick(r["home_team_shots_on_target"], r["away_team_shots_on_target"]), pick(r["away_team_shots_on_target"], r["home_team_shots_on_target"])
        posf = pick(r["home_team_possession"], r["away_team_possession"])
        ff, fa = pick(r["home_team_fouls"], r["away_team_fouls"]), pick(r["away_team_fouls"], r["home_team_fouls"])
        htf, hta = pick(r["home_team_goal_count_half_time"], r["away_team_goal_count_half_time"]), pick(r["away_team_goal_count_half_time"], r["home_team_goal_count_half_time"])
        oh, od_, oa = r["odds_ft_home_team_win"], r["odds_ft_draw"], r["odds_ft_away_team_win"]
        try:
            ph, pd_, pa = 1 / oh, 1 / od_, 1 / oa
            s = ph + pd_ + pa
            fav = (ph / s) if is_home else (pa / s)
        except Exception:
            fav = None

        # sentinela -1 (dado ausente no CSV, comum em goleadas contra seleções
        # fracas que o FootyStats não cobre em detalhe). Não deixamos -1 cru
        # na planilha (pareceria dado real negativo); usamos um placeholder
        # conservador coerente com o placar e sinalizamos para o Claude
        # revisar/ajustar antes de fechar a análise (Regra 2 do protocolo).
        missing = cf == -1
        if missing:
            print(f"⚠️  {team} x {opp} ({d.strftime('%d/%m/%Y')}): estatísticas "
                  f"ausentes no CSV (-1) além do placar. Usando placeholder "
                  f"conservador — revisar/ajustar manualmente se o jogo for "
                  f"relevante para a análise.")
            margem = gf - ga
            if margem >= 3:  # goleada
                cf, ca, yf, ya, sf, sa, sotf, sota, posf, ff, fa = 9, 1, 1, 2, 22, 3, 9, 1, 68, 9, 12
                if xgf == 0 and xga == 0:
                    xgf, xga = 3.1, 0.2
            else:
                cf, ca, yf, ya, sf, sa, sotf, sota, posf, ff, fa = 5, 4, 1, 1, 12, 10, 5, 4, 50, 10, 10
                if xgf == 0 and xga == 0:
                    xgf, xga = 1.3, 1.1

        rec = dict(
            data=d.strftime("%d/%m/%Y"), comp=comp, adv=opp, mando=mando,
            gf=int(gf), ga=int(ga), xgf=round(float(xgf), 2), xga=round(float(xga), 2),
            cf=int(cf), ca=int(ca), yf=int(yf), ya=int(ya),
            sf=int(sf), sa=int(sa), sotf=int(sotf), sota=int(sota),
            posf=int(posf), ff=int(ff), fa=int(fa), htf=int(htf), hta=int(hta),
            fav=round(fav, 3) if fav else None, missing_stats=missing,
        )
        out.append(rec)
    return out


def attach_estilo(historico, estilo_db_path=ESTILO_DB_PATH, overrides=None):
    """Preenche estilo_text/bb/pa/tr/pos/bp em cada linha do histórico usando
    o banco persistente de estilos (data/estilos_selecoes.json), com
    `overrides` (dict adversário -> tupla) tendo prioridade. Levanta erro
    listando adversários faltantes para o Claude preencher manualmente."""
    db = json.load(open(estilo_db_path, encoding="utf-8"))
    overrides = overrides or {}
    faltando = []
    for r in historico:
        if r["adv"] in overrides:
            est = overrides[r["adv"]]
            r["estilo_text"], r["bb"], r["pa"], r["tr"], r["pos"], r["bp"] = est
        elif r["adv"] in db:
            e = db[r["adv"]]
            r["estilo_text"], r["bb"], r["pa"], r["tr"], r["pos"], r["bp"] = (
                e["estilo_text"], e["bb"], e["pa"], e["tr"], e["pos"], e["bp"])
        else:
            faltando.append(r["adv"])
    if faltando:
        raise ValueError(
            "Faltam notas de estilo para: " + ", ".join(sorted(set(faltando))) +
            ". Preencha em ESTILO overrides (dict adversario -> "
            "(texto, bloco_baixo, pressao_alta, transicao, posse, bola_parada), "
            "1-5 cada) e rode attach_estilo de novo. Depois de decidir, "
            "salve no banco persistente com save_estilo_db() para reaproveitar."
        )
    return historico


def save_estilo_db(new_entries, estilo_db_path=ESTILO_DB_PATH):
    """new_entries: dict {nome_time: (texto, bb, pa, tr, pos, bp)}.
    Mescla no banco persistente para as próximas planilhas."""
    db = json.load(open(estilo_db_path, encoding="utf-8"))
    for name, (texto, bb, pa, tr, pos, bp) in new_entries.items():
        db[name] = dict(estilo_text=texto, bb=bb, pa=pa, tr=tr, pos=pos, bp=bp)
    json.dump(db, open(estilo_db_path, "w", encoding="utf-8"), indent=1, ensure_ascii=False)


# ---------------------------------------------------------------------------
# 2. CLONAGEM DE FÓRMULAS (a parte que permite redimensionar as tabelas)
# ---------------------------------------------------------------------------

def clone_formula(template_text, template_row, new_row):
    """Substitui referências de linha RELATIVAS (sem $ antes do número) que
    batem com template_row por new_row. Referências ABSOLUTAS ($linha) —
    usadas nos limites de range fixos — não são tocadas aqui (ver
    fix_range_end)."""
    pattern = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")

    def repl(m):
        col, dollar_row, row = m.groups()
        if dollar_row == "" and row == str(template_row):
            return f"{col}{new_row}"
        return m.group(0)

    return pattern.sub(repl, template_text)


def fix_range_end(text, old_end, new_end):
    """Ajusta o limite final de ranges absolutos do tipo $COL$2:$COL$old_end
    para $COL$2:$COL$new_end (usado ao redimensionar Times ou Jogos do Dia)."""
    pattern = re.compile(r"(\$[A-Z]{1,3}\$2:\$[A-Z]{1,3}\$)" + str(old_end) + r"\b")
    return pattern.sub(lambda m: m.group(1) + str(new_end), text)


def _set_formula(ws, row, col, text, is_array):
    if is_array:
        ref = f"{get_column_letter(col)}{row}"
        ws.cell(row, col, ArrayFormula(ref=ref, text=text))
    else:
        ws.cell(row, col, text)


def _capture_template(ws, row, cols):
    """Lê as fórmulas da linha-modelo `row` nas colunas `cols` (lista de
    índices). Retorna dict col -> (texto, is_array)."""
    out = {}
    for col in cols:
        v = ws.cell(row, col).value
        if v is None:
            continue
        if isinstance(v, ArrayFormula):
            out[col] = (v.text, True)
        else:
            out[col] = (v, False)
    return out


# ---------------------------------------------------------------------------
# 3. CONSTRUÇÃO DA PLANILHA
# ---------------------------------------------------------------------------

TIMES_INPUT_COLS = {
    "Time": 1, "Data": 2, "Competição": 3, "Adversário": 4, "Mando": 5,
    "Gols Pró": 6, "Gols Contra": 7, "xG Pró": 8, "xG Contra": 9,
    "Escanteios Pró": 10, "Escanteios Contra": 11, "Cartões Pró": 12, "Cartões Contra": 13,
    "Chutes Pró": 14, "Chutes Contra": 15, "Chutes Gol Pró": 16, "Chutes Gol Contra": 17,
    "Posse Pró": 18, "Faltas Time": 19, "Faltas Adv": 20, "Gols 1T Time": 21, "Gols 1T Adv": 22,
    "Favoritismo": 26, "Estilo Adv": 27, "Bloco Baixo": 28, "Pressão Alta": 29,
    "Transição": 30, "Posse/Dom": 31, "Bola Parada": 32,
}
TIMES_FORMULA_COLS = {"W": 23, "X": 24, "Y": 25, "AG": 33, "AH": 34, "AI": 35, "AK": 37}
JDD_FORMULA_COLS = list(range(15, 37))  # O..AJ
MER_FORMULA_COLS = [6, 8, 9, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21]  # F,H,I,J,L,M,O..U

MERCADOS_TEMPLATE_20 = [
    # (mercado, tipo, crit1_template, crit2_template)  {A}=TimeA {B}=TimeB
    ("BTTS Sim", "BTTS", "BTTS Sim", None),
    ("BTTS Não", "BTTS", "BTTS Não", None),
    ("CA: {FAV} DC + Under 3.5", "Combinação", "{FAVKEY} DC", "Under 3.5 | 0.84"),
    ("CA: {FAV} vence + Over 1.5", "Combinação", "{FAVKEY} vence", "Over 1.5"),
    ("CA: {FAV} vence + BTTS Não", "Combinação", "{FAVKEY} vence", "BTTS Não | 0.70"),
    ("{A} DC", "Dupla Chance", "TimeA DC", None),
    ("{B} DC", "Dupla Chance", "TimeB DC", None),
    ("Over 0.5 gols 1ºT", "Gols 1º Tempo", "Over 0.5 1T", None),
    ("Under 0.5 gols 1ºT", "Gols 1º Tempo", "Under 0.5 1T", None),
    ("{FAV} -0.5 AH (vence)", "Handicap Asiático", "{FAVKEY} AH -0.5", None),
    ("{DOG} +0.5 AH", "Handicap Asiático", "{DOGKEY} AH +0.5", None),
    ("{FAV} -1.5 AH", "Handicap Asiático", "{FAVKEY} AH -1.5", None),
    ("{A} vence", "Resultado", "TimeA vence", None),
    ("Empate", "Resultado", "Empate", None),
    ("{B} vence", "Resultado", "TimeB vence", None),
    ("Over 2.5 gols", "Total gols", "Over 2.5", None),
    ("Under 2.5 gols", "Total gols", "Under 2.5", None),
    ("Over 1.5 gols", "Total gols", "Over 1.5", None),
    ("Under 1.5 gols", "Total gols", "Under 1.5", None),
    ("Over 3.5 gols", "Total gols", "Over 3.5", None),
]


def mercados_rows_for_game(teamA, teamB, favorito, odds_and_pfont, displayA=None, displayB=None):
    """Gera as 20 linhas de Mercados para um jogo a partir do template acima.
    favorito: "A" ou "B" (qual dos dois é o favorito, define os textos das
    linhas de DC/AH combinadas).
    odds_and_pfont: lista de 20 tuplas (p_font_str, odd_mercado) NA MESMA
    ORDEM de MERCADOS_TEMPLATE_20 (pesquisar odds reais + estimar via Poisson
    onde não houver cotação direta — ver SKILL.md passo 4).
    displayA/displayB: nome em português para os rótulos de mercado (ex.:
    "França", "Marrocos"). Se omitido, usa teamA/teamB (nome do CSV) direto.
    IMPORTANTE: teamA/teamB devem ser exatamente os nomes usados em Times/
    Jogos do Dia (chave de lookup); displayA/displayB é só cosmético."""
    dA, dB = displayA or teamA, displayB or teamB
    fav_name, dog_name = (dA, dB) if favorito == "A" else (dB, dA)
    fav_key, dog_key = ("TimeA", "TimeB") if favorito == "A" else ("TimeB", "TimeA")
    jogo = f"{teamA} x {teamB}"
    rows = []
    for (mercado_t, tipo, c1_t, c2_t), (pfont, odd) in zip(MERCADOS_TEMPLATE_20, odds_and_pfont):
        mercado = mercado_t.format(A=dA, B=dB, FAV=fav_name, DOG=dog_name)
        c1 = c1_t.format(FAVKEY=fav_key, DOGKEY=dog_key) if c1_t else c1_t
        c2 = c2_t
        rows.append((jogo, mercado, tipo, c1, c2, pfont, odd))
    return rows


def build_workbook(games, template_path, output_path, data_jogo):
    """games: lista de dicts, um por partida do dia, cada um com:
        teamA, teamB           -- nomes exatamente como usados no CSV/planilha
        hist_A, hist_B         -- listas retornadas por get_historico()+attach_estilo()
        jdd_A, jdd_B           -- dict com campos de Jogos do Dia (ver SKILL.md):
            estilo_time, estilo_adv, fav, bb, pa, tr, pos, bp, ataca_fundo,
            contexto, obs
        mercados_rows          -- lista de 20 tuplas, ver mercados_rows_for_game()
        fontes                 -- dict: placar_modal, p_a, p_empate, p_b,
                                   consenso, melhor_aposta, atencao
    data_jogo: string "DD/MM/AAAA" usada na coluna Data de Jogos do Dia.

    Constrói a planilha completa (Times + Jogos do Dia + Mercados + Fontes)
    a partir do template, redimensionando as 3 tabelas Excel para o número
    exato de jogos do dia, e salva em output_path.
    """
    n_games = len(games)
    n_times_rows = n_games * 30       # 2 times x 15 jogos
    n_jdd_rows = n_games * 2          # 2 linhas por jogo (perspectiva A e B)
    n_mer_rows = n_games * 20         # 20 mercados por jogo

    wb = load_workbook(template_path)

    # ---------------- TIMES ----------------
    ws = wb["Times"]
    times_tpl = {name: ws.cell(2, col).value for name, col in TIMES_FORMULA_COLS.items()}
    old_last_row = ws.tables["Tabela15"].ref.split(":")[1]
    old_last_row_n = int(re.search(r"\d+", old_last_row).group())
    if old_last_row_n > n_times_rows + 1:
        ws.delete_rows(n_times_rows + 2, old_last_row_n - (n_times_rows + 1))
    elif old_last_row_n < n_times_rows + 1:
        ws.insert_rows(old_last_row_n + 1, (n_times_rows + 1) - old_last_row_n)

    def fix5to(text, new_n):
        return fix_range_end(text, 5, new_n)

    r = 2
    for g in games:
        for team_name, hist in ((g["teamA"], g["hist_A"]), (g["teamB"], g["hist_B"])):
            for rec in hist:
                for label, col in TIMES_INPUT_COLS.items():
                    key = {
                        "Time": None, "Data": "data", "Competição": "comp", "Adversário": "adv",
                        "Mando": "mando", "Gols Pró": "gf", "Gols Contra": "ga",
                        "xG Pró": "xgf", "xG Contra": "xga", "Escanteios Pró": "cf",
                        "Escanteios Contra": "ca", "Cartões Pró": "yf", "Cartões Contra": "ya",
                        "Chutes Pró": "sf", "Chutes Contra": "sa", "Chutes Gol Pró": "sotf",
                        "Chutes Gol Contra": "sota", "Posse Pró": "posf", "Faltas Time": "ff",
                        "Faltas Adv": "fa", "Gols 1T Time": "htf", "Gols 1T Adv": "hta",
                        "Favoritismo": "fav", "Estilo Adv": "estilo_text", "Bloco Baixo": "bb",
                        "Pressão Alta": "pa", "Transição": "tr", "Posse/Dom": "pos",
                        "Bola Parada": "bp",
                    }[label]
                    ws.cell(r, col, team_name if key is None else rec[key])
                for name, col in TIMES_FORMULA_COLS.items():
                    txt = clone_formula(times_tpl[name], 2, r)
                    if name in ("AG", "AH", "AI"):
                        txt = fix5to(txt, n_jdd_rows + 1)
                    ws.cell(r, col, txt)
                r += 1

    tbl = ws.tables["Tabela15"]
    tbl.ref = f"A1:AI{n_times_rows + 1}"
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref

    # ---------------- JOGOS DO DIA ----------------
    wsj = wb["Jogos do Dia"]
    jdd_tpl = _capture_template(wsj, 2, JDD_FORMULA_COLS)
    old_last = int(re.search(r"\d+", wsj.tables["Tabela26"].ref.split(":")[1]).group())
    if old_last > n_jdd_rows + 1:
        wsj.delete_rows(n_jdd_rows + 2, old_last - (n_jdd_rows + 1))
    elif old_last < n_jdd_rows + 1:
        wsj.insert_rows(old_last + 1, (n_jdd_rows + 1) - old_last)

    def fix61to(text, new_n):
        return fix_range_end(text, 61, new_n)

    r = 2
    for g in games:
        for team_name, adv_name, jdd in (
            (g["teamA"], g["teamB"], g["jdd_A"]), (g["teamB"], g["teamA"], g["jdd_B"])
        ):
            wsj.cell(r, 1, data_jogo)
            wsj.cell(r, 2, team_name)
            wsj.cell(r, 3, adv_name)
            wsj.cell(r, 4, jdd["estilo_time"])
            wsj.cell(r, 5, jdd["estilo_adv"])
            wsj.cell(r, 6, jdd["fav"])
            wsj.cell(r, 7, jdd["bb"])
            wsj.cell(r, 8, jdd["pa"])
            wsj.cell(r, 9, jdd["tr"])
            wsj.cell(r, 10, jdd["pos"])
            wsj.cell(r, 11, jdd["bp"])
            wsj.cell(r, 12, jdd["ataca_fundo"])
            wsj.cell(r, 13, jdd["contexto"])
            wsj.cell(r, 14, jdd["obs"])
            for col, (txt, is_arr) in jdd_tpl.items():
                new_txt = clone_formula(txt, 2, r)
                new_txt = fix61to(new_txt, n_times_rows + 1)
                _set_formula(wsj, r, col, new_txt, is_arr)
            r += 1

    tbl2 = wsj.tables["Tabela26"]
    tbl2.ref = f"A1:AH{n_jdd_rows + 1}"
    if tbl2.autoFilter is not None:
        tbl2.autoFilter.ref = tbl2.ref

    # ---------------- MERCADOS ----------------
    wsm = wb["Mercados"]
    mer_tpl = _capture_template(wsm, 2, MER_FORMULA_COLS)
    old_last = int(re.search(r"\d+", wsm.tables["Tabela37"].ref.split(":")[1]).group())
    if old_last > n_mer_rows + 1:
        wsm.delete_rows(n_mer_rows + 2, old_last - (n_mer_rows + 1))
    elif old_last < n_mer_rows + 1:
        wsm.insert_rows(old_last + 1, (n_mer_rows + 1) - old_last)

    r = 2
    for g in games:
        for row in g["mercados_rows"]:
            jogo, mercado, tipo, c1, c2, pfont, odd = row
            wsm.cell(r, 1, jogo)
            wsm.cell(r, 2, mercado)
            wsm.cell(r, 3, tipo)
            wsm.cell(r, 4, c1)
            wsm.cell(r, 5, c2)
            wsm.cell(r, 7, pfont)
            wsm.cell(r, 11, odd)
            for col, (txt, is_arr) in mer_tpl.items():
                new_txt = clone_formula(txt, 2, r)
                if col in (20, 21):  # T, U -> ranges de Jogos do Dia
                    new_txt = fix_range_end(new_txt, 5, n_jdd_rows + 1)
                _set_formula(wsm, r, col, new_txt, is_arr)
            r += 1

    tbl3 = wsm.tables["Tabela37"]
    tbl3.ref = f"A1:Q{n_mer_rows + 1}"
    if tbl3.autoFilter is not None:
        tbl3.autoFilter.ref = tbl3.ref

    # ---------------- FONTES ----------------
    wsf = wb["Fontes"]
    max_existing = wsf.max_row
    for rr in range(2, max_existing + 1):
        for c in range(1, 9):
            wsf.cell(rr, c).value = None
    for i, g in enumerate(games):
        rr = 2 + i
        f = g["fontes"]
        jogo = f"{g['teamA']} x {g['teamB']}"
        wsf.cell(rr, 1, jogo)
        wsf.cell(rr, 2, f["placar_modal"])
        wsf.cell(rr, 3, f["p_a"])
        wsf.cell(rr, 4, f["p_empate"])
        wsf.cell(rr, 5, f["p_b"])
        wsf.cell(rr, 6, f["consenso"])
        wsf.cell(rr, 7, f["melhor_aposta"])
        wsf.cell(rr, 8, f["atencao"])

    wb.save(output_path)
    return output_path
