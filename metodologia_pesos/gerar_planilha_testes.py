"""Planilha de testes visuais — Série A, todos os jogos (2023-2026).

Gera/recalcula um .xlsx com 3 abas:
  - Parâmetros: células editáveis com TODOS os parâmetros do MODELO
    (k_mando, usar_estilo, filtro_aderencia, filtro_estilo,
    filtro_favoritismo, multiplicador_dp, limite_unilateral,
    n_historico) e o limiar_edge (esse último recalcula na hora, sem
    rodar nada — é só comparação de números já calculados).
  - Jogos: todos os jogos da Série A com histórico suficiente, odds reais,
    probabilidade do MODELO e do MERCADO por mercado (Over 1.5/2.5/3.5/4.5,
    BTTS), edge, "Aposta?" e lucro — as 3 últimas são FÓRMULA VIVA do
    Excel (recalculam na hora ao mudar limiar_edge).
  - Resumo: n_apostas/taxa de acerto/lucro/ROI por mercado, também fórmula
    viva, lendo direto da aba Jogos (faixa de linhas fixa e generosa —
    `LINHA_MAX_JOGOS` — pra continuar funcionando mesmo se `n_historico`/
    filtros mudarem quantos jogos entram).

IMPORTANTE — dois tipos de parâmetro, dois fluxos diferentes:
  1. limiar_edge: muda o resultado ao vivo, o Excel recalcula sozinho.
  2. Todos os outros (k_mando, usar_estilo, filtro_aderencia,
     filtro_estilo, filtro_favoritismo, multiplicador_dp,
     limite_unilateral, n_historico): mudam a PREVISÃO do modelo em si
     (o histórico ponderado de cada time) — não dá pra recalcular só
     com fórmula sem reimplementar todo o motor de pesos no Excel (o
     que reintroduziria o mesmo tipo de bug que motivou construir esse
     motor em Python). Pra esses, é preciso rodar `computar_jogos`
     de novo — manualmente (`python3 gerar_planilha_testes.py`) ou pelo
     botão "Run main" do Excel (ver `xlwings_recalcular.py` e
     `docs/planilha_botao_recalcular.md` pro setup do botão).

Uso (linha de comando):
    python3 gerar_planilha_testes.py [caminho.xlsx]

Se o arquivo já existir, lê os parâmetros da aba Parâmetros antes de
recalcular (preserva o que o usuário editou). Se não existir, usa os
parâmetros padrão (o melhor candidato validado até 25/08/2026 — ver
docs/retrospectiva_grid_completo_2026-08-25.md).
"""
import sys
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from retrospectiva import rodar_retrospectiva

CAMINHO_PADRAO = "SerieA_testes_visuais.xlsx"

# faixa fixa de linhas que as fórmulas da aba Resumo sempre olham na aba
# Jogos — generosa o bastante pra cobrir qualquer combinação de parâmetros
# (o total de jogos da Série A 2023-2026 é 1374; linhas sobrando ficam
# em branco e não afetam COUNTIF/SUMIF).
LINHA_MAX_JOGOS = 1400

PARAMS_PADRAO = dict(
    k_mando=0.5, usar_estilo=False, filtro_aderencia=0.8,
    filtro_estilo=None, filtro_favoritismo=None,  # None = usa filtro_aderencia pros dois (padrão antigo)
    multiplicador_dp=1.5, limite_unilateral=2, n_historico=15,
    limiar_edge=0.05,
)

_CHAVES_PARAMS = (
    "k_mando", "usar_estilo", "filtro_aderencia", "filtro_estilo", "filtro_favoritismo",
    "multiplicador_dp", "limite_unilateral", "n_historico", "limiar_edge",
)

MERCADOS = [
    ("over15", "Over 1.5", "odd_over15", "prob_modelo_over15", "prob_mercado_over15", "over15_real"),
    ("over25", "Over 2.5", "odd_over25", "prob_modelo_over25", "prob_mercado_over25", "over25_real"),
    ("over35", "Over 3.5", "odd_over35", "prob_modelo_over35", "prob_mercado_over35", "over35_real"),
    ("over45", "Over 4.5", "odd_over45", "prob_modelo_over45", "prob_mercado_over45", "over45_real"),
    ("btts", "BTTS", "odd_btts_sim", "prob_modelo_btts", "prob_mercado_btts", "btts_real"),
]

FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
FUNDO_CABECALHO = PatternFill("solid", fgColor="1F4E78")
FUNDO_PARAM = PatternFill("solid", fgColor="FFF2CC")


def carregar_jogos_seriea():
    dfs = []
    for ano in ("2023", "2024", "2025"):
        df = pd.read_csv(f"data/footystats_seriea_{ano}/matches.csv")
        df["__src"] = f"m{ano}.csv"
        df = df[df["status"] == "complete"].copy()
        dfs.append(df)
    df26 = pd.read_csv("data/footystats_seriea/matches.csv")
    df26["__src"] = "m2026.csv"
    df26 = df26[df26["status"] == "complete"].copy()
    return pd.concat(dfs + [df26], ignore_index=True, sort=False)


def _valor_param(chave, valor):
    """Converte o valor bruto de uma célula da aba Parâmetros pro tipo certo
    de cada chave — usado tanto lendo de um arquivo salvo (openpyxl) quanto
    lendo do livro aberto no Excel (xlwings, mesma convenção de células)."""
    vazio = valor in (None, "", "None")
    if chave in ("k_mando", "filtro_estilo", "filtro_favoritismo"):
        return None if vazio else float(valor)
    if chave == "usar_estilo":
        return str(valor).strip().upper() in ("TRUE", "VERDADEIRO", "1")
    if chave == "n_historico":
        return PARAMS_PADRAO["n_historico"] if vazio else int(float(valor))
    return PARAMS_PADRAO[chave] if vazio else float(valor)


def ler_parametros_existentes(caminho):
    """Lê os parâmetros já salvos na aba Parâmetros de um arquivo existente."""
    wb = load_workbook(caminho, data_only=True)
    if "Parâmetros" not in wb.sheetnames:
        return dict(PARAMS_PADRAO)
    ws = wb["Parâmetros"]
    valores = {}
    for row in ws.iter_rows(min_row=3, max_col=3, values_only=False):
        nome_cel, valor_cel = row[0], row[2]
        if nome_cel.value in _CHAVES_PARAMS:
            valores[nome_cel.value] = valor_cel.value
    params = dict(PARAMS_PADRAO)
    for chave, valor in valores.items():
        params[chave] = _valor_param(chave, valor)
    return params


def montar_aba_parametros(wb, params):
    ws = wb.create_sheet("Parâmetros", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 70

    ws["A1"] = "Parâmetros do modelo — Série A"
    ws["A1"].font = Font(bold=True, size=14)

    linhas = [
        ("k_mando", params["k_mando"] if params["k_mando"] is not None else "",
         "Encolhimento de mando (0-1). Vazio = sem ajuste (None). MUDA A PREVISÃO — precisa recalcular."),
        ("usar_estilo", "TRUE" if params["usar_estilo"] else "FALSE",
         "TRUE/FALSE — usa aderência de estilo no peso. MUDA A PREVISÃO — precisa recalcular."),
        ("filtro_aderencia", params["filtro_aderencia"],
         "Mínimo de aderência (estilo/favoritismo) pro jogo histórico entrar (0-1) — usado quando filtro_estilo/"
         "filtro_favoritismo abaixo estão vazios. MUDA A PREVISÃO — precisa recalcular."),
        ("filtro_estilo", params["filtro_estilo"] if params["filtro_estilo"] is not None else "",
         "Corte de aderência SÓ de estilo, independente do de favoritismo. Vazio = usa filtro_aderencia. "
         "MUDA A PREVISÃO — precisa recalcular."),
        ("filtro_favoritismo", params["filtro_favoritismo"] if params["filtro_favoritismo"] is not None else "",
         "Corte de aderência SÓ de favoritismo, independente do de estilo. Vazio = usa filtro_aderencia. "
         "MUDA A PREVISÃO — precisa recalcular."),
        ("multiplicador_dp", params["multiplicador_dp"],
         "Corte de outlier: multiplicador do desvio-padrão. MUDA A PREVISÃO — precisa recalcular."),
        ("limite_unilateral", params["limite_unilateral"],
         "Corte de outlier: limite pra decidir unilateral/bilateral. MUDA A PREVISÃO — precisa recalcular."),
        ("n_historico", params["n_historico"],
         "Quantos jogos passados de cada time entram no cálculo (janela de histórico). MUDA A PREVISÃO — "
         "precisa recalcular."),
        ("limiar_edge", params["limiar_edge"],
         "Só conta como aposta se (prob. modelo − prob. mercado) ≥ isso. RECALCULA NA HORA, é só editar."),
    ]
    for i, (nome, valor, explicacao) in enumerate(linhas):
        r = 3 + i
        ws.cell(r, 1, nome).font = Font(bold=True)
        cel_valor = ws.cell(r, 3, valor)
        cel_valor.fill = FUNDO_PARAM
        ws.cell(r, 4, explicacao).alignment = Alignment(wrap_text=True)

    linha_edge = 3 + len(linhas) - 1  # última linha da lista = limiar_edge
    linha_ajuda = linha_edge + 2
    ws.cell(linha_ajuda, 1, "Como usar").font = Font(bold=True, size=12)
    ws.cell(linha_ajuda + 1, 1, (
        f"limiar_edge (linha {linha_edge}): edite e veja a aba Resumo recalcular na hora — é só comparação, "
        "o Excel já faz sozinho.\n\n"
        f"Todos os outros parâmetros (linhas 3-{linha_edge - 1}): mudam o histórico ponderado de cada time — "
        "não dá pra recalcular só com fórmula. Depois de editar, clique no botão \"Run main\" da aba xlwings "
        "(ver docs/planilha_botao_recalcular.md) ou rode:\n"
        "    python3 gerar_planilha_testes.py " + CAMINHO_PADRAO + "\n"
        "O script/botão lê os valores que você deixou aqui e reprocessa a aba Jogos com o novo modelo."
    )).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=linha_ajuda + 1, start_column=1, end_row=linha_ajuda + 6, end_column=4)
    return ws


def montar_aba_jogos(wb, jogos):
    ws = wb.create_sheet("Jogos")
    ws.freeze_panes = "D2"

    col = 1
    ws.cell(1, col, "Data"); col += 1
    ws.cell(1, col, "Confronto"); col += 1
    ws.cell(1, col, "Placar real"); col += 1

    inicio_mercado = {}
    for chave, titulo, *_ in MERCADOS:
        inicio_mercado[chave] = col
        for sub in ("Odd mercado", "Prob. modelo", "Prob. mercado", "Edge", "Aposta?", "Real", "Lucro"):
            c = ws.cell(1, col, f"{titulo} — {sub}")
            c.font = FONTE_CABECALHO
            c.fill = FUNDO_CABECALHO
            col += 1

    for i, jogo in enumerate(jogos):
        r = i + 2
        ws.cell(r, 1, str(jogo["data"]))
        ws.cell(r, 2, jogo["jogo"])
        ws.cell(r, 3, f"{int(jogo['gf_real'])}x{int(jogo['ga_real'])}")
        for chave, titulo, campo_odd, campo_pm, campo_pmk, campo_real in MERCADOS:
            base = inicio_mercado[chave]
            col_odd = get_column_letter(base)
            col_pm = get_column_letter(base + 1)
            col_pmk = get_column_letter(base + 2)
            col_edge = get_column_letter(base + 3)
            col_aposta = get_column_letter(base + 4)

            odd = jogo.get(campo_odd)
            prob_modelo = jogo.get(campo_pm)
            prob_mercado = jogo.get(campo_pmk)
            real = jogo.get(campo_real)

            ws.cell(r, base, odd if odd is not None else "")
            ws.cell(r, base + 1, prob_modelo if prob_modelo is not None else "")
            ws.cell(r, base + 2, prob_mercado if prob_mercado is not None else "")
            ws.cell(r, base + 3, f"={col_pm}{r}-{col_pmk}{r}" if prob_modelo is not None and prob_mercado is not None else "")
            ws.cell(r, base + 4,
                    f'=IF({col_edge}{r}="","",IF({col_edge}{r}>=Parâmetros!$C$11,"SIM",""))'
                    if prob_modelo is not None and prob_mercado is not None else "")
            ws.cell(r, base + 5, ("SIM" if real else "NÃO") if real is not None else "")
            col_real = get_column_letter(base + 5)
            col_lucro_formula = (
                f'=IF({col_aposta}{r}="SIM",IF({col_real}{r}="SIM",{col_odd}{r}-1,-1),"")'
            )
            ws.cell(r, base + 6, col_lucro_formula if odd is not None else "")

    largura = {1: 12, 2: 26, 3: 12}
    for c in range(4, col):
        largura[c] = 13
    for c, w in largura.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws, inicio_mercado, len(jogos)


def montar_aba_resumo(wb, inicio_mercado):
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    ws["A1"] = "Resumo por mercado (recalcula na hora ao mudar limiar_edge)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:E1")

    cabecalho = ["Mercado", "Apostas", "Taxa acerto", "Lucro (u)", "ROI"]
    for i, titulo in enumerate(cabecalho):
        c = ws.cell(3, i + 1, titulo)
        c.font = FONTE_CABECALHO
        c.fill = FUNDO_CABECALHO

    ultima_linha_jogos = LINHA_MAX_JOGOS  # faixa fixa e generosa, ver LINHA_MAX_JOGOS
    for i, (chave, titulo, *_resto) in enumerate(MERCADOS):
        r = 4 + i
        base = inicio_mercado[chave]
        col_aposta = get_column_letter(base + 4)
        col_lucro = get_column_letter(base + 6)
        rng_aposta = f"Jogos!{col_aposta}2:{col_aposta}{ultima_linha_jogos}"
        rng_lucro = f"Jogos!{col_lucro}2:{col_lucro}{ultima_linha_jogos}"
        rng_real = f"Jogos!{get_column_letter(base + 5)}2:{get_column_letter(base + 5)}{ultima_linha_jogos}"

        ws.cell(r, 1, titulo)
        ws.cell(r, 2, f'=COUNTIF({rng_aposta},"SIM")')
        col_n = f"B{r}"
        ws.cell(r, 3, f'=IFERROR(COUNTIFS({rng_aposta},"SIM",{rng_real},"SIM")/{col_n},"")')
        ws.cell(r, 4, f'=SUMIF({rng_aposta},"SIM",{rng_lucro})')
        ws.cell(r, 5, f'=IFERROR(D{r}/{col_n},"")')
        ws.cell(r, 3).number_format = "0.0%"
        ws.cell(r, 5).number_format = "0.0%"
    return ws


def computar_jogos(params, df=None):
    """Roda o motor de pesos (walk-forward) com os `params` dados e retorna a
    lista de jogos avaliados, pronta pra popular a aba Jogos — usado tanto
    pelo gerador de linha de comando quanto pelo botão do Excel (xlwings).

    `df`: opcional, dataset já carregado (evita reler os CSVs a cada
    recálculo quando chamado repetidamente, ex. pelo servidor do xlwings).
    """
    if df is None:
        df = carregar_jogos_seriea()
    params_modelo = {k: v for k, v in params.items() if k not in ("limiar_edge", "n_historico")}
    rel = rodar_retrospectiva(
        df, params=params_modelo, min_jogos_historico=8, min_jogos_estilo=5,
        n_historico=params["n_historico"],
    )
    print(f"{rel['n']} jogos avaliados, {rel['n_pulados']} pulados (histórico insuficiente)")
    return rel["jogos"]


def gerar_workbook(caminho, params, jogos):
    """Monta as 3 abas (Parâmetros/Jogos/Resumo) num Workbook novo e salva
    em `caminho` — usado tanto pelo gerador de linha de comando quanto pelo
    botão do Excel (`SerieA_testes_visuais.py`, xlwings)."""
    wb = Workbook()
    wb.remove(wb.active)
    montar_aba_parametros(wb, params)
    _, inicio_mercado, _ = montar_aba_jogos(wb, jogos)
    montar_aba_resumo(wb, inicio_mercado)
    wb.save(caminho)
    print(f"Salvo em {caminho}")


def gerar(caminho):
    if os.path.exists(caminho):
        params = ler_parametros_existentes(caminho)
        print(f"Lendo parâmetros existentes de {caminho}: {params}")
    else:
        params = dict(PARAMS_PADRAO)
        print(f"Arquivo novo, usando parâmetros padrão: {params}")

    jogos = computar_jogos(params)
    gerar_workbook(caminho, params, jogos)


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else CAMINHO_PADRAO
    gerar(caminho)
