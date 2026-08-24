"""Grava na planilha (template `Times`/`Jogos do Dia`) os valores calculados
pelo motor de pesos em `pesos.py` — em vez de fórmulas array/LET do Excel.

Roda como pós-processamento DEPOIS de `planilha_lib.build_workbook` já ter
montado a planilha (histórico bruto em `Times`, alvos em `Jogos do Dia`).
Substitui as duas etapas que hoje só existiam como fórmula Excel reescrita
a cada sessão (protocolo, seção "Fórmula robusta Pró/Contra"):

1. `aplicar_pesos_historico`  — recalcula Times!AG:AK (Aderência Estilo,
   Aderência Favoritismo, Peso Recência, Peso Final) usando `pesos.py`.
2. `aplicar_indicadores_pro_contra` — recalcula os 12 indicadores Pró/Contra
   de `Jogos do Dia` (média/desvio-padrão ponderados + corte de outlier).

Por serem valores Python já calculados (não fórmula), o resultado não
depende de LibreOffice/Excel avaliar LET/SUMPRODUCT corretamente — o
arquivo já sai com os números certos.
"""
from __future__ import annotations

from datetime import date, datetime

from openpyxl import load_workbook

from pesos import (
    ajuste_mando,
    calcular_pesos_historico,
    indicador_pro_contra,
)

# ---- Colunas da aba Times (histórico bruto por jogo) ----
TIMES_COL = dict(
    time="A", data="B", mando="E",
    gols_pro="F", gols_contra="G",
    escanteios_pro="J", escanteios_contra="K",
    cartoes_pro="L", cartoes_contra="M",
    chutes_pro="N", chutes_contra="O",
    chutes_gol_pro="P", chutes_gol_contra="Q",
    gols_1t_pro="U", gols_1t_contra="V",
    favoritismo="Z",
    bloco_baixo="AB", pressao_alta="AC", transicao="AD", posse="AE", bola_parada="AF",
    aderencia_estilo="AG", aderencia_favoritismo="AH", peso_recencia="AI", peso_final="AK",
)
TIMES_CAMPOS_STAT = [
    "gols_pro", "gols_contra", "cartoes_pro", "cartoes_contra",
    "escanteios_pro", "escanteios_contra", "chutes_pro", "chutes_contra",
    "chutes_gol_pro", "chutes_gol_contra", "gols_1t_pro", "gols_1t_contra",
]

# ---- Colunas da aba Jogos do Dia (alvo + indicadores por time/jogo) ----
JDD_COL = dict(
    data="A", time="B", favoritismo="F",
    bloco_baixo="G", pressao_alta="H", transicao="I", posse="J", bola_parada="K",
)
# (coluna do indicador final, coluna auxiliar _avg, campo stat correspondente em Times)
JDD_INDICADORES = [
    ("O", "BC", "gols_pro"),
    ("P", "BD", "gols_contra"),
    ("R", "BE", "cartoes_pro"),
    ("S", "BF", "cartoes_contra"),
    ("U", "BG", "escanteios_pro"),
    ("V", "BH", "escanteios_contra"),
    ("X", "BI", "chutes_pro"),
    ("Y", "BJ", "chutes_contra"),
    ("AA", "BK", "chutes_gol_pro"),
    ("AB", "BL", "chutes_gol_contra"),
    ("AD", "BM", "gols_1t_pro"),
    ("AE", "BN", "gols_1t_contra"),
]

FILTRO_ADERENCIA_PADRAO = 0.65  # mesmo corte >=65%/>=65% já usado no template-base


def _to_date(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        return datetime.strptime(valor, "%d/%m/%Y").date()
    raise TypeError(f"data em formato inesperado: {valor!r}")


def _ler_historico_times(ws_times):
    """Agrupa as linhas da aba Times por nome de time.

    Retorna dict[time] -> lista de dicts, cada um com os campos brutos do
    jogo histórico e `_row` (número da linha, para escrever de volta).
    """
    por_time = {}
    for row in ws_times.iter_rows(min_row=2):
        time = row[_col_idx(TIMES_COL["time"])].value
        if not time:
            continue
        registro = dict(_row=row[0].row)
        registro["data"] = _to_date(row[_col_idx(TIMES_COL["data"])].value)
        registro["mando"] = row[_col_idx(TIMES_COL["mando"])].value
        registro["favoritismo"] = row[_col_idx(TIMES_COL["favoritismo"])].value
        registro["notas_estilo_adv"] = [
            row[_col_idx(TIMES_COL[campo])].value
            for campo in ("bloco_baixo", "pressao_alta", "transicao", "posse", "bola_parada")
        ]
        for campo in TIMES_CAMPOS_STAT:
            registro[campo] = row[_col_idx(TIMES_COL[campo])].value
        por_time.setdefault(time, []).append(registro)
    return por_time


def _col_idx(col_letra):
    """Converte letra de coluna ('A', 'AG', ...) em índice 0-based."""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(col_letra) - 1


def _ler_alvos_jogos_do_dia(ws_jdd):
    """Retorna dict[time] -> dict(estilo_alvo, favoritismo_alvo, data_jogo, _row).

    Usa a PRIMEIRA linha de `Jogos do Dia` para cada time (mesmo critério do
    INDEX/MATCH do template: um time joga uma vez por dia nesta planilha).
    """
    alvos = {}
    for row in ws_jdd.iter_rows(min_row=2):
        time = row[_col_idx(JDD_COL["time"])].value
        if not time or time in alvos:
            continue
        estilo_alvo = [
            row[_col_idx(JDD_COL[campo])].value
            for campo in ("bloco_baixo", "pressao_alta", "transicao", "posse", "bola_parada")
        ]
        alvos[time] = dict(
            estilo_alvo=estilo_alvo,
            favoritismo_alvo=row[_col_idx(JDD_COL["favoritismo"])].value,
            data_jogo=_to_date(row[_col_idx(JDD_COL["data"])].value),
        )
    return alvos


def aplicar_pesos_historico(path, mando_alvo_por_time=None, k_mando=0.35):
    """Recalcula Times!AG:AK com o motor de `pesos.py` e grava os VALORES.

    `mando_alvo_por_time`: dict opcional {time: 'Casa'/'Fora'} — quando
    informado, aplica o ajuste de mando (shrinkage `k_mando`) igual ao
    `aplicar_ajuste_mando` do protocolo. Sem isso, só recalcula AG/AH/AI/AK
    puros (sem ajuste de mando).

    Sobrescreve o arquivo em `path` com os valores calculados. Retorna dict
    {time: historico_com_pesos} para quem quiser auditar/encadear com
    `aplicar_indicadores_pro_contra` sem reabrir o arquivo.
    """
    wb = load_workbook(path)
    ws_times = wb["Times"]
    por_time = _ler_historico_times(ws_times)
    alvos = _ler_alvos_jogos_do_dia(wb["Jogos do Dia"])

    resultado = {}
    for time, historico in por_time.items():
        alvo = alvos.get(time)
        if alvo is None:
            continue  # time sem jogo hoje em Jogos do Dia (não deveria acontecer)
        com_pesos = calcular_pesos_historico(
            historico, alvo["estilo_alvo"], alvo["favoritismo_alvo"], alvo["data_jogo"]
        )
        if mando_alvo_por_time and time in mando_alvo_por_time:
            com_pesos = ajuste_mando(com_pesos, mando_alvo_por_time[time], k=k_mando)
        for original, calculado in zip(historico, com_pesos):
            linha = original["_row"]
            ws_times.cell(linha, _col_idx(TIMES_COL["aderencia_estilo"]) + 1, calculado["aderencia_estilo"])
            ws_times.cell(linha, _col_idx(TIMES_COL["aderencia_favoritismo"]) + 1, calculado["aderencia_favoritismo"])
            ws_times.cell(linha, _col_idx(TIMES_COL["peso_recencia"]) + 1, calculado["peso_recencia"])
            ws_times.cell(linha, _col_idx(TIMES_COL["peso_final"]) + 1, calculado["peso_final"])
        resultado[time] = com_pesos

    wb.save(path)
    return resultado


def aplicar_indicadores_pro_contra(
    path,
    historico_por_time=None,
    filtro_aderencia=FILTRO_ADERENCIA_PADRAO,
    limite_unilateral=4,
    multiplicador_dp=2.5,
):
    """Recalcula os 12 indicadores Pró/Contra de `Jogos do Dia` (modelo
    validado 16/07/2026: média/desvio-padrão ponderados + corte de outlier).

    Roda DEPOIS de `aplicar_pesos_historico` (usa `peso_final` já
    calculado). `historico_por_time`: reaproveita o dict retornado por
    `aplicar_pesos_historico` para não reabrir/reler o arquivo; se omitido,
    relê `Times` do próprio `path` (assume que já tem AG/AH/AK calculados).

    Grava o valor final em O/P/R/S/U/V/X/Y/AA/AB/AD/AE e a média bruta
    (pré-corte, para auditoria) nas colunas auxiliares BC:BN — mesma
    convenção do protocolo.
    """
    wb = load_workbook(path)
    ws_times = wb["Times"]
    ws_jdd = wb["Jogos do Dia"]

    if historico_por_time is None:
        historico_por_time = _ler_historico_times(ws_times)

    # cabeçalhos das colunas auxiliares _avg (só na primeira vez)
    for _, col_avg, campo_stat in JDD_INDICADORES:
        cabecalho = ws_jdd[f"{col_avg}1"]
        if cabecalho.value is None:
            cabecalho.value = f"_avg {campo_stat}"

    relatorio = []
    for row in ws_jdd.iter_rows(min_row=2):
        time = row[_col_idx(JDD_COL["time"])].value
        if not time:
            continue
        linha = row[0].row
        historico = historico_por_time.get(time, [])
        validos = [
            j for j in historico
            if j.get("aderencia_estilo", 0) >= filtro_aderencia
            and j.get("aderencia_favoritismo", 0) >= filtro_aderencia
        ]
        pesos = [j["peso_final"] for j in validos]

        linha_relatorio = dict(time=time, row=linha, n_validos=len(validos))
        for col_ind, col_avg, campo_stat in JDD_INDICADORES:
            valores = [j.get(campo_stat) for j in validos]
            r = indicador_pro_contra(valores, pesos, limite_unilateral, multiplicador_dp)
            ws_jdd[f"{col_ind}{linha}"] = r["media_final"] if r["media_final"] is not None else 0
            ws_jdd[f"{col_avg}{linha}"] = r["media_bruta"] if r["media_bruta"] is not None else 0
            linha_relatorio[campo_stat] = r
        relatorio.append(linha_relatorio)

    # garante Parâmetros!B10/B11 documentados (não sobrescreve se já existirem)
    ws_param = wb["Parâmetros"]
    if ws_param["A10"].value is None:
        ws_param["A10"] = "Limite para corte unilateral (média)"
        ws_param["B10"] = limite_unilateral
        ws_param["C10"] = "Abaixo deste valor de média, o corte de outlier vira unilateral (só corta picos pra cima)"
    if ws_param["A11"].value is None:
        ws_param["A11"] = "Multiplicador do desvio-padrão"
        ws_param["B11"] = multiplicador_dp
        ws_param["C11"] = "limite de corte = multiplicador x desvio-padrão ponderado"

    wb.save(path)
    return relatorio
