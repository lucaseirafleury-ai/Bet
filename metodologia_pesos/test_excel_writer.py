"""Teste de integração: monta uma planilha real a partir do template (via
`planilha_lib.build_workbook`) com histórico fabricado e confirma que
`excel_writer` recalcula Times!AG:AK e os 12 indicadores Pró/Contra de
`Jogos do Dia` com valores numéricos sãos (sem depender de LibreOffice/
Excel avaliar nenhuma fórmula — os dois módulos escrevem valor pronto).
"""
import os
import tempfile

import pytest
from openpyxl import load_workbook

from excel_writer import aplicar_indicadores_pro_contra, aplicar_pesos_historico
from planilha_lib import build_workbook, mercados_rows_for_game

TEMPLATE = os.path.join(os.path.dirname(__file__), "templates", "Copa_Template_Simplificado.xlsx")


def _jogo_historico(dia, mes, ano, adv, mando, gf, ga, fav, bb, pa, tr, pos, bp):
    return dict(
        data=f"{dia:02d}/{mes:02d}/{ano}", comp="Brasileirão A", adv=adv, mando=mando,
        gf=gf, ga=ga, xgf=1.4, xga=1.1, cf=5, ca=4, yf=2, ya=3,
        sf=12, sa=10, sotf=5, sota=4, posf=50, ff=10, fa=11, htf=0, hta=0,
        fav=fav, estilo_text="", bb=bb, pa=pa, tr=tr, pos=pos, bp=bp,
    )


@pytest.fixture()
def planilha_de_teste(tmp_path):
    hist_A = [
        _jogo_historico(1, 8, 2026, "Adv1", "Casa", 2, 0, 0.60, 3, 3, 3, 3, 3),
        _jogo_historico(15, 6, 2026, "Adv2", "Fora", 1, 1, 0.55, 3, 3, 3, 3, 3),
        _jogo_historico(1, 1, 2026, "Adv3", "Casa", 0, 3, 0.20, 1, 1, 1, 1, 1),
    ]
    hist_B = [
        _jogo_historico(2, 8, 2026, "Adv4", "Fora", 1, 1, 0.40, 3, 3, 3, 3, 3),
        _jogo_historico(20, 5, 2026, "Adv5", "Casa", 2, 2, 0.45, 3, 3, 3, 3, 3),
    ]
    jdd_A = dict(estilo_time="Organizado", estilo_adv="Retranca", fav=0.60,
                 bb=3, pa=3, tr=3, pos=3, bp=3, ataca_fundo="S",
                 contexto="teste", obs="teste")
    jdd_B = dict(estilo_time="Contra-ataque", estilo_adv="Posse", fav=0.40,
                 bb=3, pa=3, tr=3, pos=3, bp=3, ataca_fundo="N",
                 contexto="teste", obs="teste")
    mer = mercados_rows_for_game(
        "TimeTesteA", "TimeTesteB", favorito="A",
        odds_and_pfont=[("50%", 2.0)] * 20,
        displayA="Time A", displayB="Time B",
    )
    games = [dict(
        teamA="TimeTesteA", teamB="TimeTesteB", hist_A=hist_A, hist_B=hist_B,
        jdd_A=jdd_A, jdd_B=jdd_B, mercados_rows=mer,
        fontes=dict(placar_modal="1-0", p_a="50%", p_empate="30%", p_b="20%",
                    consenso="teste", melhor_aposta="teste", atencao="teste"),
    )]
    out_path = str(tmp_path / "Teste.xlsx")
    build_workbook(games, template_path=TEMPLATE, output_path=out_path, data_jogo="10/08/2026")
    return out_path


def test_aplicar_pesos_historico_grava_valores_numericos(planilha_de_teste):
    resultado = aplicar_pesos_historico(planilha_de_teste)
    assert set(resultado.keys()) == {"TimeTesteA", "TimeTesteB"}
    assert len(resultado["TimeTesteA"]) == 3

    wb = load_workbook(planilha_de_teste)
    ws = wb["Times"]
    for row in ws.iter_rows(min_row=2, max_row=6):
        if row[0].value is None:
            continue
        ag, ah, ai, ak = row[32].value, row[33].value, row[34].value, row[36].value
        for valor in (ag, ah, ai, ak):
            assert isinstance(valor, (int, float))
        assert 0 <= ag <= 1
        assert 0 <= ah <= 1
        assert ai in (1.00, 0.85, 0.70, 0.50, 0.30, 0.15, 0.0)
        assert ak == pytest.approx(ag * ah * ai)


def test_pipeline_completo_gera_indicadores_pro_contra(planilha_de_teste):
    historico = aplicar_pesos_historico(planilha_de_teste)
    relatorio = aplicar_indicadores_pro_contra(planilha_de_teste, historico_por_time=historico)
    assert len(relatorio) == 2  # 2 linhas em Jogos do Dia (Time A, Time B)

    wb = load_workbook(planilha_de_teste)
    ws = wb["Jogos do Dia"]
    # coluna O = Gols Pró, coluna BC = _avg gols_pro (auditoria)
    for linha in (2, 3):
        gols_pro = ws[f"O{linha}"].value
        avg_gols_pro = ws[f"BC{linha}"].value
        assert isinstance(gols_pro, (int, float))
        assert isinstance(avg_gols_pro, (int, float))
        assert gols_pro >= 0

    # Parâmetros!B10/B11 documentados
    ws_param = wb["Parâmetros"]
    assert ws_param["B10"].value == 4
    assert ws_param["B11"].value == 2.5


def test_aplicar_indicadores_sem_historico_passado_rele_do_arquivo(planilha_de_teste):
    aplicar_pesos_historico(planilha_de_teste)
    # não passa historico_por_time -> relê Times do próprio arquivo
    relatorio = aplicar_indicadores_pro_contra(planilha_de_teste)
    assert len(relatorio) == 2
